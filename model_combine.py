import os
import math
import sys
from scipy.stats import zscore
import openpyxl
from CTkMessagebox import CTkMessagebox
from scipy.stats import trim_mean
from scipy.stats.mstats import winsorize
import numpy as np
from mathutils import Vector
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Font,Color
import datetime
from prediction_algorithm import tahmin


def model_birlestirme(model_listesi,z_score_threshold,tahmin_secenegi):
    global liste1, ikinci_konum, kolonlar
    liste_guncel = []
    liste_ana = []

    for model in model_listesi:
        liste1 = []
        wb_obj = openpyxl.load_workbook(model)

        sheet_obj = wb_obj.active
        row = sheet_obj.max_row
        column = sheet_obj.max_column

        for i in range(2, row + 1):
            liste_bos = []
            for kolon in range(1,column+1):
                cell_obj1 = sheet_obj.cell(row=i, column=kolon)
                x = str(cell_obj1.font)
                x2 = x.find('FFFF0000')
                if x2 == -1:
                    pass
                else:
                    liste_bos.append(cell_obj1.value)
            if not liste_bos:
                pass
            else:
                liste1.append(liste_bos)
            wb_obj.close()
        liste_ana.append(liste1)

    duzeltilmis_liste = []


    for nokta_index in range(len(liste_ana[0])):
        try:
            ayni_noktalar = [grup[nokta_index] for grup in liste_ana]
            kolonlar = list(zip(*ayni_noktalar))
        except:
            pass

        duzeltilmis_nokta = []
        for kolon_index, kolon in enumerate(kolonlar):

            kolon = np.array([np.nan if x is None else x for x in kolon], dtype=float)

            if kolon_index < 3:
                gecerli_indeksler = ~np.isnan(kolon)
                gecerli_degerler = kolon[gecerli_indeksler]

                if len(gecerli_degerler) > 1:  # At least 2 values are required for Z-score
                    z_skorlari = zscore(gecerli_degerler)
                    ortalama = np.nanmean(kolon)  # Calculate average ignoring NaN values
                    duzeltilmis_kolon = kolon.copy()
                    duzeltilmis_kolon[gecerli_indeksler] = [deger if abs(z) < z_score_threshold else ortalama for
                                                            deger, z in zip(gecerli_degerler, z_skorlari)]
                    duzeltilmis_nokta.append(
                        int(np.round(np.nanmean(duzeltilmis_kolon))))  # Round and convert to integer
                else:
                    duzeltilmis_nokta.append(int(np.round(
                        np.nanmean(kolon))))  # If there is not enough data, take the average directly, round and convert to integer
            else:
                duzeltilmis_nokta.append(np.nanmean(kolon))  # Take average ignoring NaN values

        duzeltilmis_liste.append(duzeltilmis_nokta)
    liste_ana2 = []

    for row in duzeltilmis_liste:
        liste_ana2.append(row)

    #List subject to change
    liste_sadece_veriler = liste_ana2

    for j in range(len(liste_ana2)):
        birinci_konum = liste_ana2[j]
        try:
            ikinci_konum = liste_ana2[j+1]
        except:
            pass

        demet1 = (birinci_konum[0],birinci_konum[1],birinci_konum[2])
        demet2 = (ikinci_konum[0],ikinci_konum[1],ikinci_konum[2])

        fark_liste = []
        fark_x = ikinci_konum[0]-birinci_konum[0]
        fark_y = ikinci_konum[1]-birinci_konum[1]
        fark_z = ikinci_konum[2]-birinci_konum[2]
        fark_liste.append(abs(fark_x))
        fark_liste.append(abs(fark_y))
        fark_liste.append(abs(fark_z))
        f = max(fark_liste)
        v1 = Vector(demet1)
        v2 = Vector(demet2)

        liste_z = []
        liste_z.append(liste_ana2[j])
        for i in range(f):
            liste_bos = []

            v_sliced = [v1.lerp(v2, (i + 1) / f)]
            liste_bos.append(int(v_sliced[0][0]))
            liste_bos.append(int(v_sliced[0][1]))
            liste_bos.append(int(v_sliced[0][2]))
            liste_z.append(liste_bos)


        for yyy in liste_z:
            liste_guncel.append(yyy)

    #elimination of repeating coordinates
    if liste_guncel[0][0] == liste_guncel[1][0] and liste_guncel[0][1] == liste_guncel[1][1] and \
            liste_guncel[0][2] == liste_guncel[ 1][2]:
        liste_guncel.pop(1)
    else:
        pass

    for t in range(len(liste_guncel)):
        try:
            if liste_guncel[t][0] == liste_guncel[t+1][0] and liste_guncel[t][1] == liste_guncel[t+1][1] and liste_guncel[t][2] == liste_guncel[t+1][2]:
                liste_guncel.pop(t)
        except:
            pass

    #Adding x,y,z and labels to the beginning of the list

    wb_obj2 = openpyxl.load_workbook(model_listesi[0])
    sheet_obj = wb_obj2.active
    column = sheet_obj.max_column
    liste_bos_2 = []
    for kolon in range(1, column + 1):
        cell_obj1 = sheet_obj.cell(row=1, column=kolon)
        liste_bos_2.append(cell_obj1.value)

    liste_guncel.insert(0,liste_bos_2)


    liste_sadece_veriler.insert(0, liste_bos_2)

    listeyu = []
    for gg in range(3,len(liste_sadece_veriler[0])):
        liste_yu_alt = []
        for gt in range(len(liste_sadece_veriler)):
            liste_yu_alt.append(liste_sadece_veriler[gt][gg])

        listeyu.append(liste_yu_alt)

    #Excel file intermediate values are empty and calculated here
    #create excel file and print final model

    wb = Workbook()
    ws = wb.active

    s1 = 1

    for satir in liste_guncel:
        s2 = 1
        for satir2 in satir:
            ws.cell(s1, s2, satir2)
            if len(satir) > 3:
                ws.cell(s1, s2).font = Font(color="FFFF0000")  # To distinguish control coordinates, they are colored red (X)
            else:
                pass
            s2 = s2 + 1
        s1 = s1 + 1


    klasor_yolu = os.path.dirname(model_listesi[0])

    tarih_damgasi = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    son_excel = r"{}/combined_model{}.xlsx".format(klasor_yolu,tarih_damgasi)
    wb.save(son_excel)

    tahmin(listeyu,tahmin_secenegi,son_excel)
    CTkMessagebox(title="Info", message="Model Merge Completed!!!", width=400, height=200)

    wb_obj2.close()


def model_birlestirme2(model_listesi,mesafe_siniri,tahmin_secenegi):

    global liste1, ikinci_konum, liste_diger
    liste_guncel = []
    liste_ana = []

    #Determination of the main model
    ana_model_liste = []
    for ana_model in model_listesi:
        x = str(ana_model)
        x2 = x.find('primary')
        if x2 == -1:
            pass
        else:
            ana_model_liste.append(ana_model)

    if not ana_model_liste:
        CTkMessagebox(title="Error", message="Main Model Missing!!!\nName the main model as 'primary'", icon="cancel", width=400, height=200)
        sys.exit()
    else:
        pass


    model = ana_model_liste[0]#main model

    liste1 = []
    wb_obj_ana = openpyxl.load_workbook(model)
    sheet_obj = wb_obj_ana.active
    row = sheet_obj.max_row
    column = sheet_obj.max_column
    for i in range(2, row + 1):
        liste_bos = []
        for kolon in range(1, column + 1):
            cell_obj1 = sheet_obj.cell(row=i, column=kolon)
            x = str(cell_obj1.font)
            x2 = x.find('FFFF0000')
            if x2 == -1:
                pass
            else:
                liste_bos.append(cell_obj1.value)
        if not liste_bos:
            pass
        else:
            liste1.append(liste_bos)
    liste_ana.append(liste1)

    for j in range(len(liste1)):
        birinci_konum = liste1[j]
        try:
            ikinci_konum = liste1[j + 1]
        except:
            pass

        demet1 = (birinci_konum[0], birinci_konum[1], birinci_konum[2])
        demet2 = (ikinci_konum[0], ikinci_konum[1], ikinci_konum[2])

        fark_liste = []
        fark_x = ikinci_konum[0] - birinci_konum[0]
        fark_y = ikinci_konum[1] - birinci_konum[1]
        fark_z = ikinci_konum[2] - birinci_konum[2]
        fark_liste.append(abs(fark_x))
        fark_liste.append(abs(fark_y))
        fark_liste.append(abs(fark_z))
        f = max(fark_liste)
        v1 = Vector(demet1)
        v2 = Vector(demet2)

        liste_z = []
        liste_z.append(liste1[j])
        for i in range(f):
            liste_bos = []

            v_sliced = [v1.lerp(v2, (i + 1) / f)]
            liste_bos.append(int(v_sliced[0][0]))
            liste_bos.append(int(v_sliced[0][1]))
            liste_bos.append(int(v_sliced[0][2]))
            liste_z.append(liste_bos)

        for yyy in liste_z:
            liste_guncel.append(yyy)

    ref_excel_liste = liste_guncel

    for diger_listeler in range(1, len(model_listesi)):
        liste_diger = []
        wb_obj = openpyxl.load_workbook(model_listesi[diger_listeler])

        sheet_obj = wb_obj.active
        row = sheet_obj.max_row
        column = sheet_obj.max_column

        for i in range(2, row + 1):
            liste_bos = []
            for kolon in range(1, column + 1):
                cell_obj1 = sheet_obj.cell(row=i, column=kolon)
                x = str(cell_obj1.font)
                x2 = x.find('FFFF0000')
                if x2 == -1:
                    pass
                else:
                    liste_bos.append(cell_obj1.value)
            if not liste_bos:
                pass
            else:
                liste_diger.append(liste_bos)

            if not liste_bos:
                pass
            else:
                koordinat = liste_bos

                en_kucuk_mesafe = float('inf')
                en_yakin_index = -1  # Variable to be assigned to the index of the closest point
                x_yeni, y_yeni, z_yeni, *degerler_yeni = koordinat
                for i, alt_liste in enumerate(ref_excel_liste):
                    x, y, z, *degerler_mevcut = alt_liste
                    # Calculate Euclidean distance
                    mesafe = math.sqrt((x - x_yeni) ** 2 + (y - y_yeni) ** 2 + (z - z_yeni) ** 2)
                    # If a smaller distance is found, update
                    if mesafe < en_kucuk_mesafe:
                        en_kucuk_mesafe = mesafe
                        en_yakin_index = i
                # If distance is greater than limit, do not merge
                if en_kucuk_mesafe > mesafe_siniri:
                    break#return ref_excel_liste  # Return without changing the list
                # If closest point is found, combine with arithmetic mean
                x, y, z, *degerler_mevcut = ref_excel_liste[en_yakin_index]
                x = (x + x_yeni) / 2
                y = (y + y_yeni) / 2
                z = (z + z_yeni) / 2
                # Take the arithmetic average by combining the current values and the new values
                yeni_degerler = []
                max_uzunluk = max(len(degerler_mevcut), len(degerler_yeni))
                for i in range(max_uzunluk):
                    mevcut_deger = degerler_mevcut[i] if i < len(degerler_mevcut) else None
                    yeni_deger = degerler_yeni[i] if i < len(degerler_yeni) else None
                    if mevcut_deger is not None and yeni_deger is not None:
                        yeni_degerler.append((mevcut_deger + yeni_deger) / 2)
                    elif mevcut_deger is not None:
                        yeni_degerler.append(mevcut_deger)
                    elif yeni_deger is not None:
                        yeni_degerler.append(yeni_deger)
                # Add updated data to main list
                ref_excel_liste[en_yakin_index] = [x, y, z] + yeni_degerler
        wb_obj.close()


    #split ref excel list into columns and re-expand

    liste_yeniden = []
    liste_guncel2 = []
    for yy in ref_excel_liste:
        if len(yy) > 3:
            liste_yeniden.append(yy)
        else:
            pass
    for j in range(len(liste_yeniden)):
        birinci_konum = liste_yeniden[j]
        try:
            ikinci_konum = liste_yeniden[j + 1]
        except:
            pass

        demet1 = (birinci_konum[0], birinci_konum[1], birinci_konum[2])
        demet2 = (ikinci_konum[0], ikinci_konum[1], ikinci_konum[2])

        fark_liste = []
        fark_x = ikinci_konum[0] - birinci_konum[0]
        fark_y = ikinci_konum[1] - birinci_konum[1]
        fark_z = ikinci_konum[2] - birinci_konum[2]
        fark_liste.append(abs(fark_x))
        fark_liste.append(abs(fark_y))
        fark_liste.append(abs(fark_z))
        f = max(fark_liste)
        v1 = Vector(demet1)
        v2 = Vector(demet2)

        liste_z = []
        liste_z.append(liste_yeniden[j])
        for i in range(int(f)):
            liste_bos = []

            v_sliced = [v1.lerp(v2, (i + 1) / f)]
            liste_bos.append(int(v_sliced[0][0]))
            liste_bos.append(int(v_sliced[0][1]))
            liste_bos.append(int(v_sliced[0][2]))
            liste_z.append(liste_bos)

        for yyy in liste_z:
            liste_guncel2.append(yyy)
    ref_excel_liste2 = liste_guncel2
    wb_obj2 = openpyxl.load_workbook(model)
    sheet_obj = wb_obj2.active
    column = sheet_obj.max_column
    liste_bos_2 = []
    for kolon in range(1, column + 1):
        cell_obj1 = sheet_obj.cell(row=1, column=kolon)
        liste_bos_2.append(cell_obj1.value)

    ref_excel_liste2.insert(0, liste_bos_2)

    liste_sadece_veriler = liste1

    liste_sadece_veriler.insert(0, liste_bos_2)

    wb = Workbook()
    ws = wb.active

    s1 = 1

    #deletion of repeating elements
    for satir_guncellenmemis in range(len(ref_excel_liste2)):

        # Compare all elements by converting them to float
        try:
            if list(map(float, ref_excel_liste2[satir_guncellenmemis])) == list(map(float, ref_excel_liste2[satir_guncellenmemis+1])):
                ref_excel_liste2.pop(satir_guncellenmemis+1)
            else:
                pass
        except:
            pass

    for satir in ref_excel_liste2:
        s2 = 1
        for satir2 in satir:
            ws.cell(s1, s2, satir2)
            if len(satir) > 3:
                ws.cell(s1, s2).font = Font(
                    color="FFFF0000")  # To distinguish control coordinates, they are colored red (X)
            else:
                pass
            s2 = s2 + 1
        s1 = s1 + 1

    klasor_yolu = os.path.dirname(model)

    tarih_damgasi = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    son_excel = r"{}/combined_model_timeline_independent{}.xlsx".format(klasor_yolu, tarih_damgasi)
    wb.save(son_excel)  # excel file intermediate values are empty



    # Obtaining control points from ref_excel_list

    sadece_kontrol_noktalari = []
    for kontrol_noktasi in ref_excel_liste2:
        if len(kontrol_noktasi) > 3:
            sadece_kontrol_noktalari.append(kontrol_noktasi)
        else:
            pass

    listeyu2 = []
    for gg in range(3, len(sadece_kontrol_noktalari[0])):
        liste_yu_alt2 = []
        for gt in range(len(sadece_kontrol_noktalari)):
            liste_yu_alt2.append(sadece_kontrol_noktalari[gt][gg])
        listeyu2.append(liste_yu_alt2)

    #only measurement data(listeyu2)

    tahmin(listeyu2, tahmin_secenegi, son_excel)
    CTkMessagebox(title="Info", message="Model Merge Completed!!!", width=400, height=200)
    wb_obj_ana.close()
    wb_obj2.close()









