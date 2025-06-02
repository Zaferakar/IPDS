import sys
import openpyxl
from scipy.optimize import brent
from list_split_combine import *
from exponential_and_linear import ustel_tahmin, lineer_tahmin, verileri_genislet, oldugu_gibi_birak
from CTkMessagebox import CTkMessagebox

def tahmin(kolonlar_veri,tahmin_secenegi,excel_konumu_ara_degerler_bos):

    for ilk_eleman_cikart in kolonlar_veri:#To make an increase estimate, the first label is deleted
        ilk_eleman_cikart.pop(0)

    global alt_liste
    wb_obj = openpyxl.load_workbook(excel_konumu_ara_degerler_bos)

    sheet_obj = wb_obj.active
    row = sheet_obj.max_row
    column = sheet_obj.max_column

    kolonlar = []
    for j in range(4, column + 1):
        liste_bos = []
        for i in range(1, row + 1):
            cell_obj = sheet_obj.cell(row=i, column=j)
            liste_bos.append(cell_obj.value)
        liste_bos.pop(0)#remove column label value from list
        kolonlar.append(liste_bos)

    sutun_tercih_sayac = 0

    for sutunlar in kolonlar:
        parcali_listeler = alt_listelere_parcala(sutunlar)
        u_l_tahmin_icin_index = 0
        #columns are processed separately
        parcali_listeleri_tutulmasi = []
        parcali_listeleri_tutulmasi_2 = []
        u_l_tahmin_icin_index_listesi = []
        for parcali in parcali_listeler:
            #manipulation position
            try:
                if tahmin_secenegi[sutun_tercih_sayac][0] == "u":
                    gelen_liste = ustel_tahmin(len(parcali)-1,parcali[0],parcali[-1])
                    gelen_liste[0] = parcali[0]
                    gelen_liste[-1] = parcali[-1]
                    parcali_listeleri_tutulmasi.append(gelen_liste)
                    alt_liste = alt_listeleri_birlestir(parcali_listeleri_tutulmasi)
                elif tahmin_secenegi[sutun_tercih_sayac][0] == "L":
                    gelen_liste = lineer_tahmin(len(parcali) - 1, parcali[0], parcali[-1])
                    gelen_liste[0] = parcali[0]
                    gelen_liste[-1] = parcali[-1]
                    parcali_listeleri_tutulmasi.append(gelen_liste)
                    alt_liste = alt_listeleri_birlestir(parcali_listeleri_tutulmasi)
                elif tahmin_secenegi[sutun_tercih_sayac][0] == "G":
                    gelen_liste = verileri_genislet(len(parcali) - 1, parcali[0], parcali[-1])
                    gelen_liste[0] = parcali[0]
                    gelen_liste[-1] = parcali[-1]
                    parcali_listeleri_tutulmasi.append(gelen_liste)
                    alt_liste = alt_listeleri_birlestir(parcali_listeleri_tutulmasi)
                elif tahmin_secenegi[sutun_tercih_sayac][0] == "u+L":
                    tahmin_icin_kolonlara_ayir = alt_listelere_parcala(kolonlar_veri[sutun_tercih_sayac])

                    for alt_alt_liste in tahmin_icin_kolonlara_ayir:#It connects to the Excel file containing the raw data and determines the period numbers.
                        ilk_deger = alt_alt_liste[0]
                        son_deger = alt_alt_liste[-1]
                        if ilk_deger == 0 and son_deger > 0:  # If the first value is zero and the last value is positive, it assumes the smallest positive value as zero to avoid errors
                            ilk_deger = 1e-10
                        elif ilk_deger == 0 and son_deger < 0:  # If the first value is zero and the last value is negative, it assumes the smallest negative value as zero to avoid errors
                            ilk_deger = -1e-10
                        else:
                            pass
                        periyot = len(alt_alt_liste)-1
                        karar_degeri = ((son_deger / ilk_deger) * (1 / periyot))
                        u_l_tahmin_icin_index_listesi.append(karar_degeri)
                    if u_l_tahmin_icin_index_listesi[u_l_tahmin_icin_index] > 3.5:
                        gelen_liste = ustel_tahmin(len(parcali) - 1, parcali[0], parcali[-1])
                        gelen_liste[0] = parcali[0]
                        gelen_liste[-1] = parcali[-1]
                        parcali_listeleri_tutulmasi_2.append(gelen_liste)
                        alt_liste = alt_listeleri_birlestir(parcali_listeleri_tutulmasi_2)
                    elif u_l_tahmin_icin_index_listesi[u_l_tahmin_icin_index] <= 3.5:
                        gelen_liste = lineer_tahmin(len(parcali) - 1, parcali[0], parcali[-1])
                        gelen_liste[0] = parcali[0]
                        gelen_liste[-1] = parcali[-1]
                        parcali_listeleri_tutulmasi_2.append(gelen_liste)
                    alt_liste = alt_listeleri_birlestir(parcali_listeleri_tutulmasi_2)
                    u_l_tahmin_icin_index = u_l_tahmin_icin_index + 1
                elif tahmin_secenegi[sutun_tercih_sayac][0] == "tahmin_yok":
                    gelen_liste = oldugu_gibi_birak(len(parcali) - 1, parcali[0], parcali[-1])
                    gelen_liste[0] = parcali[0]
                    gelen_liste[-1] = parcali[-1]
                    parcali_listeleri_tutulmasi.append(gelen_liste)
                    alt_liste = alt_listeleri_birlestir(parcali_listeleri_tutulmasi)
            except:
                    CTkMessagebox(title="Error",
                                  message="First Row Value Cannot Be Blank in Excel Data File,\nTraining Stopped, Update Value!!!",
                                  icon="cancel", width=400, height=200)
                    sys.exit()
        try:
            satir_sayac = 2
            for satir in alt_liste:
                sheet_obj.cell(satir_sayac,sutun_tercih_sayac+4,satir)
                wb_obj.save(excel_konumu_ara_degerler_bos)
                satir_sayac = satir_sayac + 1
        except:
            sys.exit()
        sutun_tercih_sayac = sutun_tercih_sayac + 1

    wb_obj.close()








