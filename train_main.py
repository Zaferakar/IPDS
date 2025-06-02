import threading
import xlsxwriter
from PIL import Image
import os
from natsort import natsorted
from mathutils import Vector
from openpyxl import Workbook,load_workbook
from CTkMessagebox import CTkMessagebox
from openpyxl.styles import Font,Color

from prediction_algorithm import *

#tahmin_kumesi = [[],[],[],... ](prediction set)
#excel_veriler = A file containing data entered into an Excel file
#klasor_eğitim = folder location with photos
#egitim_model_ismi = name of the model to be trained

def model_egitim(tahmin_kumesi,excel_veriler,egitim_model_ismi,klasor_egitim):

    #Writing coordinates between main coordinates to excel file
    excel_uzanti = '{}/{}.xlsx'.format(klasor_egitim,egitim_model_ismi)
    workbook = xlsxwriter.Workbook(excel_uzanti)
    worksheet = workbook.add_worksheet()

    dosyalar = os.listdir(klasor_egitim)
    fotolar = []
    for foto in dosyalar:

        if foto.endswith(".png"):
            fotolar.append("{}/{}".format(klasor_egitim, foto))
        elif foto.endswith(".jpg"):
            fotolar.append("{}/{}".format(klasor_egitim, foto))
        elif foto.endswith(".jpeg"):
            fotolar.append("{}/{}".format(klasor_egitim, foto))
        else:
            pass
    siralanmis_klasor_icerigi = natsorted(fotolar)

    listee = []
    def rgb_kod(im):
        im2 = Image.open(im, 'r')
        pixel_values = list(im2.getdata())
        RGB_x = []
        RGB_y = []
        RGB_z = []
        for i in pixel_values:
            RGB_x.append(i[0])
            RGB_y.append(i[1])
            RGB_z.append(i[2])
        son_x = round(sum(RGB_x) / len(RGB_x))
        son_y = round(sum(RGB_y) / len(RGB_y))
        son_z = round(sum(RGB_z) / len(RGB_z))

        liste_bos = []

        liste_bos.append(son_x)
        liste_bos.append(son_y)
        liste_bos.append(son_z)
        listee.append(liste_bos)
        return son_x, son_y, son_z
    #Titles
    worksheet.write("A1", "X")
    worksheet.write("B1", "Y")
    worksheet.write("C1", "Z")

    for foto_konum in siralanmis_klasor_icerigi:
        rgb_kod(foto_konum)

    #Since the initial value is disabled in vectorization, we write it ourselves

    worksheet.write("A2", listee[0][0])
    worksheet.write("B2", listee[0][1])
    worksheet.write("C2", listee[0][2])
    #Processing vector elements into excel file
    parca_sayisi_liste = []
    satir2 = 3

    for j in range(len(listee)-1):
        birinci_konum = listee[j]
        ikinci_konum = listee[j+1]

        demet1 = (birinci_konum[0],birinci_konum[1],birinci_konum[2])
        demet2 = (ikinci_konum[0],ikinci_konum[1],ikinci_konum[2])
        fark_liste = []
        fark_x = ikinci_konum[0]-birinci_konum[0]
        fark_y = ikinci_konum[1]-birinci_konum[1]
        fark_z = ikinci_konum[2]-birinci_konum[2]
        fark_liste.append(abs(fark_x))
        fark_liste.append(abs(fark_y))
        fark_liste.append(abs(fark_z))
        f = max(fark_liste)
        v1 = Vector(demet1)
        v2 = Vector(demet2)
        parca_sayisi_liste.append(f)

        for i in range(f):
            v_sliced = [v1.lerp(v2, (i + 1) / f)]
            worksheet.write("A{}".format(satir2), int(v_sliced[0][0]))
            worksheet.write("B{}".format(satir2), int(v_sliced[0][1]))
            worksheet.write("C{}".format(satir2), int(v_sliced[0][2]))

            satir2 = satir2 + 1
    parca_sayisi_liste[0] = parca_sayisi_liste[0]+2
    do = 1
    satirlar = [1, 2, parca_sayisi_liste[0]]  #control coordinates !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
    satir_baslangic = parca_sayisi_liste[0]

    for jo in range(len(parca_sayisi_liste) - 1):
        satir_baslangic = satir_baslangic + parca_sayisi_liste[do]
        do = do + 1
        satirlar.append(satir_baslangic)
    workbook.close()


    wb_obj = openpyxl.load_workbook(excel_veriler)
    sheet_obj = wb_obj.active
    row = sheet_obj.max_row#total number of rows
    column = sheet_obj.max_column#total number of columns

    kolonlar = []
    for j in range(1, column + 1):
        liste_bos = []
        for i in range(1, row + 1):
            cell_obj = sheet_obj.cell(row=i, column=j)


            liste_bos.append(cell_obj.value)
        kolonlar.append(liste_bos)

    wb_obj2 = openpyxl.load_workbook(excel_uzanti)#Re-open the same excel file to prevent the error
    sheet_obj2 = wb_obj2.active
    #printing values according to control coordinates

    #The number of rows in the excel file and the number of cropped photos must be equal, if it gives an error, there is a mismatch
    try:
        for kolon_sayisi in range(column):
            for satir_sayisi in range(row):
                sheet_obj2.cell(satirlar[satir_sayisi],kolon_sayisi+4,kolonlar[kolon_sayisi][satir_sayisi])
                sheet_obj2.cell(satirlar[satir_sayisi],kolon_sayisi+4,kolonlar[kolon_sayisi][satir_sayisi]).font = Font(color="FFFF0000")#Control coordinates are colored red to distinguish them

                sheet_obj2.cell(satirlar[satir_sayisi], kolon_sayisi + 1).font = Font(color="FFFF0000")#Control coordinates are colored red to distinguish them(X)
                sheet_obj2.cell(satirlar[satir_sayisi], kolon_sayisi + 2).font = Font(color="FFFF0000")#Control coordinates are colored red to distinguish them(Y)
                sheet_obj2.cell(satirlar[satir_sayisi], kolon_sayisi + 3).font = Font(color="FFFF0000")#Control coordinates are colored red to distinguish them(Z)

                wb_obj2.save(excel_uzanti)
        wb_obj2.close()
    except:
        CTkMessagebox(title="Error", message="Excel Data File Contains More Data Than Cropped Photos,\n Crop Missing Photos or Edit Excel Data !!!",
                      icon="cancel", width=400, height=200)

    threading.Thread(target=tahmin,args=(kolonlar,tahmin_kumesi,excel_uzanti,)).start()






