import threading
from contextlib import suppress
from pathlib import Path
import model_compare
from functions import *
from mark_image_coordinates import koordinat_isaretleme
from train_main import *
import customtkinter
import mplcyberpunk
from matplotlib.axes import Axes
from matplotlib.backends.backend_tkagg import (
    FigureCanvasTkAgg, NavigationToolbar2Tk)
from matplotlib.figure import Figure
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from prediction_main import *
from model_combine import model_birlestirme, model_birlestirme2

customtkinter.set_appearance_mode("dark")

class App(customtkinter.CTk):

    def __init__(self):
        super().__init__()
        cikarim_deneme_liste = []

        self.geometry("1245x670")
        self.title("AKAR")
        # tain layer(interface)
        def egitim():
            self.egitim = customtkinter.CTkFrame(self, width=1600, height=600)
            self.egitim.grid(column=0, row=1, columnspan=3,rowspan=3, pady=10, padx=10)
            #first frame
            self.canvas = customtkinter.CTkCanvas(self.egitim,width=1504, height=705)
            self.canvas.grid(row=2,column=0,columnspan=10)
            self.vsb = customtkinter.CTkScrollbar(self.egitim, orientation="vertical", command=self.canvas.yview)
            self.vsb.grid(column=10, row=2,  sticky=N+S+W)
            self.canvas.configure(yscrollcommand=self.vsb.set)
            self.hsb = customtkinter.CTkScrollbar(self.egitim, orientation="horizontal", command=self.canvas.xview)
            self.hsb.place(relx=1, rely=1, anchor="se", relwidth=1)
            self.canvas.configure(xscrollcommand=self.hsb.set,bg="grey17")
            self.main_frame = customtkinter.CTkFrame(self.canvas)
            self.canvas.create_window((0, 0), window=self.main_frame, anchor="nw")
            self.main_frame.update_idletasks()
            self.canvas.configure(scrollregion=self.canvas.bbox("all"))

            klasor_liste_son = []
            klasor_egitim = []
            metot_liste = []

            def cati():
                global klasor_konum
                global klasor_konumu_dosya_sayisi_liste
                klasor_konumu_dosya_sayisi_liste = klasor_secme()# sorted location of photos, folder location, excel file location


                def renk(i):
                    self.foto_kirpma = customtkinter.CTkLabel(self.main_frame, width=150, text="Image Saved",
                                                              fg_color="green", corner_radius=8)
                    self.foto_kirpma.grid(row=i + 2, column=3, pady=5, padx=2)

                if klasor_konumu_dosya_sayisi_liste[0] == "":
                    klasor_liste_son.append("")
                else:
                    klasor_liste_son.append(klasor_konumu_dosya_sayisi_liste)


                #constantly updated frame
                if klasor_liste_son[-1] == "" :
                    pass
                else:#image processing layer
                    #create folder
                    try:
                        klasor_deger = klasor_olusturma(klasor_konumu_dosya_sayisi_liste[-2])
                        klasor_egitim.append(klasor_deger)
                        self.klasor_sec = customtkinter.CTkLabel(self.egitim, width=520,
                                                                  text="Scanning Folder...",
                                                                  fg_color="red4",corner_radius=8)
                        self.klasor_sec.grid(row=1, column=1, pady=5, padx=2)
                        self.canvas = customtkinter.CTkCanvas(self.egitim, width=1504, height=705)
                        self.canvas.grid(row=2, column=0, columnspan=10)
                        self.vsb = customtkinter.CTkScrollbar(self.egitim, orientation="vertical", command=self.canvas.yview)
                        self.vsb.grid(column=10, row=2, sticky=N + S + W)
                        self.canvas.configure(yscrollcommand=self.vsb.set)
                        self.hsb = customtkinter.CTkScrollbar(self.egitim, orientation="horizontal", command=self.canvas.xview)
                        self.hsb.place(relx=1, rely=1, anchor="se", relwidth=1)
                        self.canvas.configure(xscrollcommand=self.hsb.set, bg="grey17")

                        self.main_frame = customtkinter.CTkFrame(self.canvas)
                        self.canvas.create_window((0, 0), window=self.main_frame, anchor="nw")
                        self.main_frame.update_idletasks()
                        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

                        self.bos = customtkinter.CTkLabel(master=self.main_frame,corner_radius=8,width=30,text="",height=70,fg_color="dark slate gray")
                        self.bos.grid(row=0,column=0,pady=5,rowspan=2)

                        self.foto_ismi_label = customtkinter.CTkLabel(master=self.main_frame, text="Image Name",
                                                                      fg_color="dark slate gray", corner_radius=8,
                                                                      width=400,height=70)
                        self.foto_ismi_label.grid(row=0,rowspan=2, column=1, pady=5)



                        self.foto_kirp_label = customtkinter.CTkLabel(master=self.main_frame, text="Crop Image", width=200,fg_color="DodgerBlue4",
                                                                       corner_radius=8,height=70)
                        self.foto_kirp_label.grid(row=0,rowspan=2,  column=2)

                        self.foto_kirp_label = customtkinter.CTkLabel(master=self.main_frame, text="Clipping Control",
                                                                      width=154, fg_color="green",
                                                                      corner_radius=8,height=70)
                        self.foto_kirp_label.grid(row=0,rowspan=2,  column=3)

                        a = 0

                        for i in range(len(klasor_konumu_dosya_sayisi_liste)-2):#folder content
                            foto = klasor_konumu_dosya_sayisi_liste[i]

                            if klasor_konumu_dosya_sayisi_liste[i].endswith(".png") or klasor_konumu_dosya_sayisi_liste[i].endswith(".jpg") or klasor_konumu_dosya_sayisi_liste[i].endswith(".jpeg"):
                                self.numara_label = customtkinter.CTkLabel(self.main_frame, width=30, text="{}".format(i + 1),
                                                                           fg_color="darkblue", corner_radius=8)
                                self.numara_label.grid(row=i + 2, column=0, pady=5, padx=2)
                                #content names label
                                self.klasor_sec2 = customtkinter.CTkLabel(self.main_frame, width=400,
                                                                          text="{}".format(klasor_konumu_dosya_sayisi_liste[i]), fg_color="dark slate gray",corner_radius=8)
                                self.klasor_sec2.grid(row=i+2, column=1, pady=5, padx=2)

                                self.foto_kirpma = customtkinter.CTkButton(self.main_frame, width=200,text="Crop Image",fg_color="DodgerBlue4", command=lambda i = foto,b = a: [koordinat_isaretleme(i,klasor_deger),renk(b)])
                                self.foto_kirpma.grid(row=i+2,column=2,pady=5, padx=2)

                                self.foto_kirpma = customtkinter.CTkLabel(self.main_frame, width=150, text="Image Not Cropped!",
                                                                           fg_color="red4",corner_radius=8)
                                self.foto_kirpma.grid(row=i + 2, column=3, pady=5, padx=2)
                                a = a + 1

                        excel_fonksiyon_degerleri = excel_okuma(klasor_konumu_dosya_sayisi_liste[-1],klasor_konumu_dosya_sayisi_liste[-2])

                        sutun_renkleri = ["grey40","grey80"]
                        sutun_renkleri_son = []
                        for sutun in range(excel_fonksiyon_degerleri[0]):
                            for sutun2 in sutun_renkleri:
                                sutun_renkleri_son.append(sutun2)
                        for sutun_kadar_liste in range(excel_fonksiyon_degerleri[0]):#Creating empty sublists as many as the number of columns
                            bos_liste = ["tahmin_yok"]
                            metot_liste.append(bos_liste)

                        def tahmin_belirleme(sutun,string):#Matching the data entered in excel according to column indexes

                            if not metot_liste[sutun]:
                                metot_liste[sutun].append(string)

                            else:
                                metot_liste[sutun].clear()
                                metot_liste[sutun].append(string)
                            return metot_liste


                        kolon_kaydirma = 5
                        excel_sutun = 0
                        kolon_baslangic = 4


                        for degerler in range(excel_fonksiyon_degerleri[0]):#column
                            excel_satir = 0

                            self.kontrol_str = customtkinter.StringVar(value="off")

                            self.ustel_kontrol = customtkinter.CTkRadioButton(self.main_frame, text=":E",width=10,variable=self.kontrol_str,value="1",command=lambda kolon = excel_sutun: tahmin_belirleme(kolon,"u")
                                                                               )
                            self.ustel_kontrol.grid(row=0, column=kolon_baslangic+kolon_kaydirma, pady=5,
                                                    padx=2)


                            self.lineer_kontrol = customtkinter.CTkRadioButton(self.main_frame, text=":L",width=10,variable=self.kontrol_str,command=lambda kolon = excel_sutun : tahmin_belirleme(kolon,"L"),
                                                                               value="2", )
                            self.lineer_kontrol.grid(row=0, column=kolon_baslangic+1+kolon_kaydirma, pady=5,
                                                      padx=2)


                            self.ustel_lineer_kontrol = customtkinter.CTkRadioButton(self.main_frame, text=":E+L",width=10,variable=self.kontrol_str,command=lambda kolon = excel_sutun : tahmin_belirleme(kolon,"u+L"),
                                                                               value="3")
                            self.ustel_lineer_kontrol.grid(row=0, column=kolon_baslangic+2+kolon_kaydirma, pady=5,
                                                      padx=2)


                            self.genislet_kontrol = customtkinter.CTkRadioButton(self.main_frame, text=":Ex",width=10,variable=self.kontrol_str,command=lambda kolon = excel_sutun : tahmin_belirleme(kolon,"G"),
                                                                               value="4")
                            self.genislet_kontrol.grid(row=0, column=kolon_baslangic+3+kolon_kaydirma, pady=5,
                                                      padx=2)

                            for satir in range(excel_fonksiyon_degerleri[1]):#line

                                self.klasor_sec2 = customtkinter.CTkLabel(self.main_frame, width=200,text=excel_fonksiyon_degerleri[2][excel_sutun][excel_satir],
                                                                          fg_color=sutun_renkleri_son[degerler],text_color="black", corner_radius=8)
                                self.klasor_sec2.grid(row=satir+1 , column=kolon_baslangic+kolon_kaydirma,columnspan = 6, pady=5, padx=2)

                                excel_satir = excel_satir + 1
                            kolon_baslangic = kolon_baslangic + 1
                            excel_sutun = excel_sutun + 1
                            kolon_kaydirma = kolon_kaydirma+ 5

                        self.klasor_sec.destroy()
                        self.main_frame.update_idletasks()
                        self.canvas.configure(scrollregion=self.canvas.bbox("all"))
                    except:
                        pass

            #read me window
            def beni_oku():
                self.beni_oku_pencere = customtkinter.CTkToplevel()
                self.beni_oku_pencere.attributes("-topmost", True)
                self.beni_oku_pencere.geometry("800x568")
                self.beni_oku_pencere.title("Train Guidelines")

                self.kaydirilabiliralan = customtkinter.CTkScrollableFrame(master=self.beni_oku_pencere,width=773,height=550)
                self.kaydirilabiliralan.grid(padx=2,pady=2,row=0,column=0)

                self.bilgi1 = customtkinter.CTkTextbox(master=self.kaydirilabiliralan,width=760,height=130)
                self.bilgi1.grid(row=0,column=0)

                self.cubuk = customtkinter.CTkLabel(master=self.kaydirilabiliralan,width=760,height=5,fg_color="grey",text="")
                self.cubuk.grid(row=2,column=0,pady=20)

                self.bilgi2 = customtkinter.CTkTextbox(master=self.kaydirilabiliralan, width=760, height=130)
                self.bilgi2.grid(row=3, column=0)

                self.bilgi3 = customtkinter.CTkTextbox(master=self.kaydirilabiliralan, width=760, height=150)
                self.bilgi3.grid(row=6, column=0)

                self.bilgi4 = customtkinter.CTkTextbox(master=self.kaydirilabiliralan, width=760, height=370)
                self.bilgi4.grid(row=9, column=0)

                self.bilgi5 = customtkinter.CTkTextbox(master=self.kaydirilabiliralan, width=760, height=150)
                self.bilgi5.grid(row=12, column=0)

                self.bilgi6 = customtkinter.CTkTextbox(master=self.kaydirilabiliralan, width=760, height=120)
                self.bilgi6.grid(row=15, column=0)

                self.bilgi7 = customtkinter.CTkTextbox(master=self.kaydirilabiliralan, width=760, height=95)
                self.bilgi7.grid(row=18, column=0)

                self.bilgi8 = customtkinter.CTkTextbox(master=self.kaydirilabiliralan, width=760, height=125)
                self.bilgi8.grid(row=21, column=0)

                self.cubuk1 = customtkinter.CTkLabel(master=self.kaydirilabiliralan, width=760, height=5,
                                                    fg_color="grey", text="")
                self.cubuk1.grid(row=2, column=0, pady=20)

                self.cubuk2 = customtkinter.CTkLabel(master=self.kaydirilabiliralan, width=760, height=5,
                                                    fg_color="grey", text="")
                self.cubuk2.grid(row=5, column=0, pady=20)

                self.cubuk3 = customtkinter.CTkLabel(master=self.kaydirilabiliralan, width=760, height=5,
                                                    fg_color="grey", text="")
                self.cubuk3.grid(row=8, column=0, pady=20)

                self.cubuk4 = customtkinter.CTkLabel(master=self.kaydirilabiliralan, width=760, height=5,
                                                    fg_color="grey", text="")
                self.cubuk4.grid(row=11, column=0, pady=20)

                self.cubuk5 = customtkinter.CTkLabel(master=self.kaydirilabiliralan, width=760, height=5,
                                                    fg_color="grey", text="")
                self.cubuk5.grid(row=14, column=0, pady=20)

                self.cubuk6 = customtkinter.CTkLabel(master=self.kaydirilabiliralan, width=760, height=5,
                                                    fg_color="grey", text="")
                self.cubuk6.grid(row=17, column=0, pady=20)

                self.cubuk7 = customtkinter.CTkLabel(master=self.kaydirilabiliralan, width=760, height=5,
                                                    fg_color="grey", text="")
                self.cubuk7.grid(row=20, column=0, pady=20)

                self.cubuk8 = customtkinter.CTkLabel(master=self.kaydirilabiliralan, width=760, height=5,
                                                     fg_color="grey", text="")
                self.cubuk8.grid(row=23, column=0, pady=20)

                self.foto1 = customtkinter.CTkImage(light_image=Image.open(Path(__file__).with_name("photos")/"a.png"),size=(760,300))
                self.foto_label1 = customtkinter.CTkLabel(master=self.kaydirilabiliralan,image=self.foto1,text="",corner_radius=8)
                self.foto_label1.grid(row=1,column=0,pady=5)

                self.foto2 = customtkinter.CTkImage(light_image=Image.open(Path(__file__).with_name("photos")/"b.png"),
                                                   size=(760, 350))
                self.foto_label2 = customtkinter.CTkLabel(master=self.kaydirilabiliralan, image=self.foto2,text="")
                self.foto_label2.grid(row=4, column=0,pady=5)

                self.foto3 = customtkinter.CTkImage(light_image=Image.open(Path(__file__).with_name("photos")/"c.png"),
                                                   size=(760, 180))
                self.foto_label3 = customtkinter.CTkLabel(master=self.kaydirilabiliralan, image=self.foto3,text="")
                self.foto_label3.grid(row=7, column=0)

                self.foto4 = customtkinter.CTkImage(light_image=Image.open(Path(__file__).with_name("photos")/"d.png"),
                                                    size=(760, 180))
                self.foto_label4 = customtkinter.CTkLabel(master=self.kaydirilabiliralan, image=self.foto4, text="")
                self.foto_label4.grid(row=10, column=0)

                self.foto5 = customtkinter.CTkImage(
                    light_image=Image.open(Path(__file__).with_name("photos")/"e.png"),
                    size=(760, 180))
                self.foto_label5 = customtkinter.CTkLabel(master=self.kaydirilabiliralan, image=self.foto5, text="")
                self.foto_label5.grid(row=13, column=0)

                self.foto6 = customtkinter.CTkImage(
                    light_image=Image.open(Path(__file__).with_name("photos")/"f.png"),
                    size=(760, 180))
                self.foto_label6 = customtkinter.CTkLabel(master=self.kaydirilabiliralan, image=self.foto6, text="")
                self.foto_label6.grid(row=16, column=0)

                self.foto7 = customtkinter.CTkImage(
                    light_image=Image.open(Path(__file__).with_name("photos")/"g.png"),
                    size=(760, 180))
                self.foto_label7 = customtkinter.CTkLabel(master=self.kaydirilabiliralan, image=self.foto7, text="")
                self.foto_label7.grid(row=19, column=0)

                self.foto8 = customtkinter.CTkImage(
                    light_image=Image.open(Path(__file__).with_name("photos")/"h.png"),
                    size=(760, 280))
                self.foto_label8 = customtkinter.CTkLabel(master=self.kaydirilabiliralan, image=self.foto8, text="")
                self.foto_label8.grid(row=22, column=0)

                self.bilgi8.insert("0.0", "Important Information for Train 3:\n \n"
                                          "Model Weight Location:\n"
                                          "The created model weight file is written in excel file format and a different folder is created under the dataset folder. The folder is named according to the training time of the model. Re-training can be done with different prediction models by changing the model name in the same folder location. As long as the model name is changed, the weights are saved in different excel files. ")

                self.bilgi7.insert("0.0","Important Information for Train 2:\n \n"
                                         "Missing/Incomplete Intermediate Values (None Value):\n"
                                         "It is not mandatory to enter data for each intermediate image. Null values can be estimated in accordance with the selected estimation model.")


                self.bilgi6.insert("0.0","Important Information for Train 1:\n \n"
                                         "First or Last Value is Missing/Incomplete (None Value):\n"
                                         "The initial value of the columns cannot be missing. If the final value is missing, the algorithm generates estimates of intermediate values assuming the final value. If the values are unknown, it is recommended to start training by re-obtaining the data. The accuracy and frequency of the data are important for the accuracy of the estimates.")

                self.bilgi5.insert("0.0","Naming the Model and Starting Training:\n \n"
                                         "After cropping the images and selecting the prediction preferences, the model name is entered and then the training is started. There is no action to be taken after the initialization.\n\n"
                                         "A folder is created under the selected folder and the model weights are saved as an Excel file in this location. This Excel file should be used for inference after training. You can delete the cropped images.")

                self.bilgi4.insert("0.0","Prediction Preferences:\n \n"
                                   "Each entered value is used as a control point. As the control points increase, the accuracy of the model increases. Four basic options are provided for the estimation of intermediate data: E (Exponential), L (Linear), E+L (Exponential + Linear) and Ex (Expansion).\n \n"
                                         "Exponential (E): Makes intermediate estimates assuming that all data follow an exponential increase-decrease model over time.\n \n"
                                         "Linear (L): Makes predictions assuming that all data fit a linear increase-decrease model over time.\n \n"
                                         "Exponential + Linear (E+L): It assumes that the data follows an exponential increase-decrease model in some time periods and a linear increase-decrease model in some time periods, and makes an estimate by looking at the difference between the data.\n \n"
                                         "Expand (Ex): Expands data over time periods to estimate the continuation of definite phenomena where the data is not numerical. (For example; Exhibition of vitality or phenotype, etc.)\n \n"
                                         "Warning!!! If no option is selected, no prediction is made and assumptions are made based on control points during inference. Since each column is a separate data set, each selection covers only that column.")

                self.bilgi3.insert("0.0","Cropping Images:\n \n"
                                   "The location of the images and Excel data file is selected with the training button. After selection, the images and Excel file are read by the program. If the Excel file is not read or the images are not loaded, check the folder location and try again. There is a crop button next to each image and when cropped, the checkbox becomes active. Even if the images have been cropped before, run the crop tool and close it again without doing anything. In this way, the images will be included in the subfolder for processing the model training. ")

                self.bilgi2.insert("0.0", "Excel Data Set Preparation:\n \n"
                                          "The Excel file must be in the same folder location as the visuals. There must not be more than one Excel file in the same folder location. Each column in the Excel file is considered as a separate data set. The first row in each column is the name of the obtained data set. The timeline must be filled from top to bottom starting from the second row. Each cell from the second row must correspond to the order of the visual in the folder (1st Cell (Row): Data set name, 2nd Cell (Row) --> 1st Visual, 3rd Cell (Row) --> 2nd Visual, ...).")
                self.bilgi1.insert("0.0","Preparing Image Dataset:\n \n"
                                         "Images should be taken at certain periodic intervals. There is a cropping tool in the program, you do not need to crop them separately. If you are going to apply any process to cropped images, click on the cropping button and close the window without applying any process. If the cropping tool does not work, the images are not included in the analysis. The images should be in the same folder and the images should be named and sorted according to the timeline. (For example; 1.jpg, 2.jpg, 3.jpg, 4.jpg, .... or a.jpg, b.jpg, c.jpg, d.jpg, ... )")


                self.bilgi1.configure(state="disable")
                self.bilgi2.configure(state="disable")
                self.bilgi3.configure(state="disable")

                self.bilgi4.configure(state="disable")
                self.bilgi5.configure(state="disable")
                self.bilgi6.configure(state="disable")
                self.bilgi7.configure(state="disable")
                self.bilgi8.configure(state="disable")


            self.beni_oku = customtkinter.CTkButton(self.egitim, text=" Directive ",fg_color="red4", width=60,command=beni_oku)
            self.beni_oku.grid(row=1, column=0, pady=5, padx=2)

            self.klasor_sec = customtkinter.CTkButton(self.egitim, width=281,
                                                      text="Select Image Dataset Folder ", fg_color="grey25",
                                                      command=lambda :threading.Thread(target=cati).start())
            self.klasor_sec.grid(row=1, column=1, pady=5, padx=2)

            # Model Name
            def model_egitim_cati():
                egitim_model_ismi = self.model_ismi.get()
                if egitim_model_ismi == "":
                    def egitim_uyari():
                        CTkMessagebox(title="Error", message="Training Failed to Start!!!\nEnter Model Name!!!",
                                  icon="cancel", width=400, height=200)
                    threading.Thread(target=egitim_uyari).start()
                else:
                    def egitim_baslat():
                        CTkMessagebox(title="Info", message="Training Has Started!!!",width=400,height=200)
                        model_egitim(metot_liste,klasor_konumu_dosya_sayisi_liste[-1],egitim_model_ismi,klasor_egitim[-1])#Parameters required for training
                    threading.Thread(target=egitim_baslat).start()


            self.model_ismi = customtkinter.CTkEntry(self.egitim,width=281,placeholder_text="Enter Model Name",fg_color="grey25")
            self.model_ismi.grid(row=1, column=2, pady=5)

            # train button
            self.egitim_butonu = customtkinter.CTkButton(self.egitim, width=281,
                                                         text="Train the Model", fg_color="grey25",command= lambda: threading.Thread(target=model_egitim_cati).start() )
            self.egitim_butonu.grid(row=1, column=3, pady=5, padx=2)

            self.main_frame.update_idletasks()
            self.canvas.configure(scrollregion=self.canvas.bbox("all"))

            #model merge(top window)
            metot_liste2 = []
            def modelbirlestir():
                secili_modeller = model_secme()
                liste_isimler = []

                try:
                    wb_obj = openpyxl.load_workbook(secili_modeller[0])
                    sheet_obj = wb_obj.active
                    column = sheet_obj.max_column
                    for hucre in range(4,column+1):
                        cell_obj1 = sheet_obj.cell(row=1, column=hucre)
                        liste_isimler.append(cell_obj1.value)

                    for ll in range(len(liste_isimler)):
                        metot_liste2.append(["tahmin_yok"])


                    def tahmin_belirleme2(sutun,
                                         string):  # Matching the data entered in excel according to column indexes

                        if not metot_liste2[sutun]:
                            metot_liste2[sutun].append(string)

                        else:
                            metot_liste2[sutun].clear()
                            metot_liste2[sutun].append(string)
                        return metot_liste2

                    self.ust_pencere = customtkinter.CTkToplevel()
                    self.ust_pencere.attributes("-topmost", True)
                    self.ust_pencere.geometry("700x520")
                    self.ust_pencere.title("Model Merge")

                    self.kaydirilabilir_alan1 = customtkinter.CTkScrollableFrame(self.ust_pencere,width=670,height=500)
                    self.kaydirilabilir_alan1.grid(row=0,column=0,pady=5,padx=5)

                    satir = 0
                    for degerler in range(len(liste_isimler)):  # column

                        self.kontrol_str = customtkinter.StringVar(value="off")
                        self.veri_seti_ismi = customtkinter.CTkLabel(self.kaydirilabilir_alan1,text="{}".format(liste_isimler[degerler]),width=250,corner_radius=8,fg_color="green")
                        self.veri_seti_ismi.grid(row=degerler, column=0)


                        self.ustel_kontrol = customtkinter.CTkRadioButton(self.kaydirilabilir_alan1, text=":E", width=100,
                                                                          variable=self.kontrol_str, value="1",command=lambda kolon = satir: tahmin_belirleme2(kolon,"u"))
                        self.ustel_kontrol.grid(row=degerler, column=1, pady=5,
                                                padx=2)

                        self.ustel_kontrol = customtkinter.CTkRadioButton(self.kaydirilabilir_alan1, text=":L", width=100,
                                                                          variable=self.kontrol_str, value="2",command=lambda kolon = satir: tahmin_belirleme2(kolon,"L"))
                        self.ustel_kontrol.grid(row=degerler, column=2, pady=5,
                                                padx=2)

                        self.ustel_kontrol = customtkinter.CTkRadioButton(self.kaydirilabilir_alan1, text=":E+L", width=100,command=lambda kolon = satir: tahmin_belirleme2(kolon,"u+L"),
                                                                          variable=self.kontrol_str, value="3")
                        self.ustel_kontrol.grid(row=degerler, column=3, pady=5,
                                                padx=2)

                        self.ustel_kontrol = customtkinter.CTkRadioButton(self.kaydirilabilir_alan1, text=":Ex", width=100,command=lambda kolon = satir: tahmin_belirleme2(kolon,"G"),
                                                                          variable=self.kontrol_str, value="4")
                        self.ustel_kontrol.grid(row=degerler, column=4, pady=5,
                                                padx=2)
                        satir = satir + 1

                    value_listesi = [0]
                    def slider_event(value):
                        value_listesi.append(value)
                        self.model_birlestir_butonu = customtkinter.CTkLabel(self.kaydirilabilir_alan1,
                                                                             text="{}".format(value_listesi[-1]), width=100)
                        self.model_birlestir_butonu.grid(row=satir + 1, column=2,columnspan=3)

                        return value_listesi[-1]#Winsorized Average Threshold

                    value_listesi2 = [0]

                    def slider_event2(value):
                        value_listesi2.append(value)
                        self.mesafe_goster = customtkinter.CTkLabel(self.kaydirilabilir_alan1,
                                                                             text="{} unit(s)".format(value_listesi2[-1]),
                                                                             width=100)
                        self.mesafe_goster.grid(row=satir + 3, column=2, columnspan=3)

                        return value_listesi2[-1]  #Distance Threshold


                    self.ort_esik = customtkinter.CTkLabel(self.kaydirilabilir_alan1,
                                                                         text="Z-Score:", width=100)
                    self.ort_esik.grid(row=satir + 1, column=0)

                    self.mesafe_esik = customtkinter.CTkLabel(self.kaydirilabilir_alan1,
                                                           text="Distance Threshold:", width=100)
                    self.mesafe_esik.grid(row=satir + 3, column=0)

                    self.baslangic_aset = customtkinter.CTkLabel(self.kaydirilabilir_alan1,
                                                                         text="0",
                                                                         width=100)
                    self.baslangic_aset.grid(row=satir + 1, column=2, columnspan=3)

                    self.baslangic_aset2 = customtkinter.CTkLabel(self.kaydirilabilir_alan1,
                                                                         text="0 unit(s)",
                                                                         width=100)
                    self.baslangic_aset2.grid(row=satir + 3, column=2, columnspan=3)

                    self.slider = customtkinter.CTkSlider(self.kaydirilabilir_alan1, from_=-10, to=10,
                                                          command=slider_event, width=200, number_of_steps=80)
                    self.slider.grid(row=satir + 1, column=1, columnspan=2, pady=20)
                    self.slider.set(0)

                    self.slider2 = customtkinter.CTkSlider(self.kaydirilabilir_alan1, from_=0, to=25,
                                                          command=slider_event2, width=200, number_of_steps=50)
                    self.slider2.grid(row=satir + 3, column=1, columnspan=2, pady=20)
                    self.slider2.set(0)

                    self.model_birlestir_butonu = customtkinter.CTkButton(self.kaydirilabilir_alan1,text="Combine Models with Time Alignment",width=670,command=lambda : threading.Thread(target=model_birlestirme,args=(secili_modeller,value_listesi[-1],metot_liste2,)).start())
                    self.model_birlestir_butonu.grid(row=satir+2,column=0,columnspan = 5)

                    self.model_birlestir_butonu2 = customtkinter.CTkButton(self.kaydirilabilir_alan1,
                                                                          text="Merge Models Directly", width=670,
                                                                          command=lambda: threading.Thread(
                                                                              target=model_birlestirme2, args=(
                                                                              secili_modeller, value_listesi2[-1],
                                                                              metot_liste2,)).start())
                    self.model_birlestir_butonu2.grid(row=satir + 4, column=0, columnspan=5)

                    self.bilgi_model = customtkinter.CTkTextbox(master=self.kaydirilabilir_alan1, width=670, height=870)
                    self.bilgi_model.grid(row=satir+5,column=0,columnspan = 5,pady = 10)

                    self.bilgi_model.insert("0.0", "Combine Models with Time Alignment:\n \n"
                                                   "In order to combine models, each model must contain an equal number of control points. To avoid disrupting the time flow, combine only models that are running in parallel. Otherwise, incompatible models will emerge and cause incorrect results."
                                                "Z-Score: A statistical measure that describes how far a data point is from the mean of a data set, measured in standard deviations."
                                                   " This is done to limit the impact of anomalous data on computation and model assembly . "
                                
                                                   "If you set it to 0, all data will be directly included in the calculation without any abnormal data detection.\n \n \n"
                                                   "Merge Models Directly:\n \n"
                                                   "Models with different timelines can be merged. When merging models, the control points that show the closest phenotype and are under the distance threshold are merged."
                                                   "1 main model named 'primary.xlsx' and at least 1 side model should be merged. Other models will be transferred to the 'primary.xlsx' file. Name the model that shows the most correct behavior as 'primary.xlsx'. \n \n"
                                                   "Distance Threshold:\n"
                                                   "The distance threshold can take a value between 0 and 25 units. As the unit distance increases, the similarity of the joined points decreases and the probability of getting erroneous results increases."
                                                   " As the unit distance decreases, the sharpness increases, but the data loss also increases significantly.\n \n \n"
                                                   "Caution !!!\n"
                                                   "Since real world values will change after model merging, the estimated intermediate values of the old models are not taken into account. After merging the models, the intermediate values must be estimated again..\n \n \n"
                                                   "Prediction Preferences:\n"
                                                   
                                   "Each entered value is used as a control point. As the control points increase, the accuracy of the model increases. Four basic options are provided for the estimation of intermediate data: E(Exponential), L(Linear), E+L(Exponential+Linear) and Ex(Expansion).\n \n"
                                         "Exponential (E): Makes intermediate estimates assuming that all data follow an exponential increase-decrease model over time..\n \n"
                                         "Linear (L): Makes predictions assuming that all data fits a linear increase-decrease model over time.\n \n"
                                         "Exponential + Linear (E+L): It assumes that the data follows an exponential increase-decrease model in some time periods and a linear increase-decrease model in some time periods, and makes an estimate by looking at the difference between the data.\n \n"
                                         "Expand (Ex): Expands data over time periods to estimate the continuation of definite phenomena where the data is not numerical (e.g., the display of vitality or phenotype, etc.).\n \n"
                                         "Caution !!!\nIf no option is selected, no estimation is made for intermediate values, but models are combined and assumptions are made according to control points during inference. Since each column is a separate data set, each selection covers only that column.")

                    self.bilgi_model.configure(state="disable")

                except:
                    pass

            self.model_birlestirme = customtkinter.CTkButton(self.egitim, width=281,
                                                             text="Model Merge ", fg_color="grey25",command=modelbirlestir
                                                             )
            self.model_birlestirme.grid(row=1, column=4, pady=5, padx=2)

        # Prediction layer(interface)
        def cikarim():
            secili_model_listesi = []

            self.cikarim = customtkinter.CTkFrame(self, fg_color="gray25")
            self.cikarim.grid(column=0, row=1, columnspan=2, pady=10, padx=10)

            self.cikarim_veri_giris = customtkinter.CTkScrollableFrame(self.cikarim, width=340, height=570)
            self.cikarim_veri_giris.grid(column=0, row=0, pady=10, padx=10)

            self.kaydir = customtkinter.CTkScrollableFrame(self.cikarim, width=800, height=570)

            self.kaydir.grid(column=10, row=0, pady=10, padx=10)

            def yonerge_cikarim_penceresi():#predict guide(top window)
                self.yonerge = customtkinter.CTkToplevel()
                self.yonerge.attributes("-topmost", True)
                self.yonerge.geometry("800x568")
                self.yonerge.title("Predict Guide")

                self.kaydirilabiliralan1 = customtkinter.CTkScrollableFrame(master=self.yonerge, width=773,
                                                                           height=550)
                self.kaydirilabiliralan1.grid(padx=2, pady=2, row=0, column=0)

                self.bilgi1 = customtkinter.CTkTextbox(master=self.kaydirilabiliralan1, width=760, height=160)
                self.bilgi1.grid(row=0, column=0)

                self.bilgi2 = customtkinter.CTkTextbox(master=self.kaydirilabiliralan1, width=760, height=90)
                self.bilgi2.grid(row=3, column=0)

                self.bilgi3 = customtkinter.CTkTextbox(master=self.kaydirilabiliralan1, width=760, height=130)
                self.bilgi3.grid(row=6, column=0)

                self.bilgi4 = customtkinter.CTkTextbox(master=self.kaydirilabiliralan1, width=760, height=320)
                self.bilgi4.grid(row=9, column=0)

                self.bilgi5 = customtkinter.CTkTextbox(master=self.kaydirilabiliralan1, width=760, height=370)
                self.bilgi5.grid(row=12, column=0)

                self.cubuk = customtkinter.CTkLabel(master=self.kaydirilabiliralan1, width=760, height=5,
                                                    fg_color="grey", text="")
                self.cubuk.grid(row=2, column=0, pady=20)

                self.cubuk1 = customtkinter.CTkLabel(master=self.kaydirilabiliralan1, width=760, height=5,
                                                     fg_color="grey", text="")
                self.cubuk1.grid(row=2, column=0, pady=20)

                self.cubuk2 = customtkinter.CTkLabel(master=self.kaydirilabiliralan1, width=760, height=5,
                                                     fg_color="grey", text="")
                self.cubuk2.grid(row=5, column=0, pady=20)

                self.cubuk3 = customtkinter.CTkLabel(master=self.kaydirilabiliralan1, width=760, height=5,
                                                     fg_color="grey", text="")
                self.cubuk3.grid(row=8, column=0, pady=20)

                self.cubuk4 = customtkinter.CTkLabel(master=self.kaydirilabiliralan1, width=760, height=5,
                                                     fg_color="grey", text="")
                self.cubuk4.grid(row=11, column=0, pady=20)

                self.cubuk5 = customtkinter.CTkLabel(master=self.kaydirilabiliralan1, width=760, height=5,
                                                     fg_color="grey", text="")
                self.cubuk5.grid(row=14, column=0, pady=20)

                self.foto2 = customtkinter.CTkImage(light_image=Image.open(Path(__file__).with_name("photos")/"i.png"),
                                                    size=(760, 440))
                self.foto_label2 = customtkinter.CTkLabel(master=self.kaydirilabiliralan1, image=self.foto2, text="")
                self.foto_label2.grid(row=4, column=0, pady=5)

                self.foto3 = customtkinter.CTkImage(light_image=Image.open(Path(__file__).with_name("photos")/"j.png"),
                                                    size=(760, 400))
                self.foto_label3 = customtkinter.CTkLabel(master=self.kaydirilabiliralan1, image=self.foto3, text="")
                self.foto_label3.grid(row=7, column=0)

                self.foto4 = customtkinter.CTkImage(
                    light_image=Image.open(Path(__file__).with_name("photos")/"k.png"),
                    size=(760, 450))
                self.foto_label4 = customtkinter.CTkLabel(master=self.kaydirilabiliralan1, image=self.foto4, text="")
                self.foto_label4.grid(row=10, column=0)

                self.foto5 = customtkinter.CTkImage(
                    light_image=Image.open(Path(__file__).with_name("photos")/"l.png"),
                    size=(760, 260))
                self.foto_label5 = customtkinter.CTkLabel(master=self.kaydirilabiliralan1, image=self.foto5, text="")
                self.foto_label5.grid(row=13, column=0)


                self.bilgi5.insert("0.0", "Viewing Prediction Results:\n \n"
                                          "Forecast results are not automatically recorded. You can copy the forecast results and transfer them to another document.\n \n"
                                          "AKAR Algorithm:\n"
                                          "The algorithm matches the input data to the most similar point on the model. After matching, the similarity between the two points is calculated with the Weighted Scoring formula and the estimated value range is determined.\n \n"
                                          "If intermediate values are not estimated during training, the algorithm finds similarity with the closest control point and estimates the value. The data similarity rate (Accuracy Rate) is also communicated to the user. The closer this rate is to 100, the more accurate the estimate is.\n"
                                          "Formation of Two Different Points:\n"
                                          "If the AKAR algorithm detects two points separately on the graph, the input data is identical to two different points. As a result, two different calculations are made, and the user can base the calculation he wants on the accuracy rates of the calculation results.\n \n"
                                          "Random Forest Regression:\n"
                                          "Random Forest Regression Function is used to estimate the numerical data in the Data Set. The estimated numerical value is shown in the graph.\n \n"
                                          "Artificial Neural Network (Softmax Activation Function):\n"
                                          "It is used to classify and predict non-numeric values in the data set. The predicted value is shown in the graph.   ")

                self.bilgi4.insert("0.0", "Predict Graph:\n \n"
                                          "Line Graph:\n"
                                          "The line graph represents the trained model. Perpendicular Lines: Perpendicular and parallel lines represent periods (one line for each image). Orange Circle: Shows the result obtained using the Random Forest Regression Function for numerical data and the Softmax Activation Function (Artificial Neural Network) for verbal data. Green Circle: Shows the result obtained using the AKAR Algorithm. Time Line: The time line is parallel to the order of the images in the data set. The distortions in the time line are related to the training phase.\n \n"
                                          "Saving the Graph: The graph can be viewed and saved via the options pane at the bottom left. ")

                self.bilgi3.insert("0.0", "Making a Prediction:\n \n"
                                          "After the prediction button is activated, the prediction can be started. After clicking the prediction start button, the control light next to the button will turn orange. After the prediction process is completed, the graphs and the predicted values are displayed on the side. The screen is updated again for each View.")
                self.bilgi2.insert("0.0", "Cropping Images:\n \n"
                                          "The exact same method is applied as in the training phase. You need to mark the part of the image to be scanned with the cropping tool. The extraction button will not be active until the area of the image to be scanned is determined.")

                self.bilgi1.insert("0.0", "Post-Training Prediction:\n \n"
                                          "After the training is completed, the created excel file will be used for prediction. Images are selected in a folder in bulk or individually. You cannot select the excel data file before the images are selected. If you forget to select the excel data file, the prediction will not start. You can select more than one excel data file. If you select more than one excel data file, the prediction is made separately for each file.  ")

                self.bilgi1.configure(state="disable")
                self.bilgi2.configure(state="disable")
                self.bilgi3.configure(state="disable")

                self.bilgi4.configure(state="disable")
                self.bilgi5.configure(state="disable")
            def cikarim_goruntuler_al_klasor(secim):
                global siralanmis_goruntu_klasor_icerigi, fotolar_liste_cikarim

                if secim == "tekil_gorsel":
                    try:
                        fotolar_liste_cikarim = gorsel_secme()
                        gecici_klasor = klasor_olusturma(fotolar_liste_cikarim[-1])
                        """print(gecici_klasor)"""
                    except:
                        with suppress(Exception):#prevents error if folder is not selected
                            cikarim()

                elif secim == "klasor":
                    try:
                        fotolar_liste_cikarim = klasor_secme()
                        gecici_klasor = klasor_olusturma(fotolar_liste_cikarim[-2])#temporary folder location
                        """print(gecici_klasor)"""
                    except:
                        with suppress(Exception):#prevents error if folder is not selected
                            cikarim()
                else:
                    pass

                sadece_fotolar = []

                for foto in fotolar_liste_cikarim:
                    if foto.endswith(".png"):
                        sadece_fotolar.append(foto)
                    elif foto.endswith(".jpg"):
                        sadece_fotolar.append(foto)
                    elif foto.endswith(".jpeg"):
                        sadece_fotolar.append(foto)
                    else:
                        pass
                #reorder images
                siralanmis_goruntu_klasor_icerigi = natsorted(sadece_fotolar)

                #just get photo names
                liste_parcali = []
                for i in siralanmis_goruntu_klasor_icerigi:
                    parca = re.split("/|\|//",i)
                    liste_parcali.append(parca[-1])
                if siralanmis_goruntu_klasor_icerigi == []:
                    CTkMessagebox(title="Error", message="Folder is empty\nimages not found!!!",
                                  icon="cancel", width=400, height=200)
                else:
                    numara = 1
                    #recreate the table
                    try:
                        self.cikarim_veri_giris.destroy()
                    except:
                        pass
                    self.cikarim_veri_giris = customtkinter.CTkScrollableFrame(self.cikarim, width=340, height=570)
                    self.cikarim_veri_giris.grid(column=0, row=0, pady=10, padx=10)
                    def cikarim_baslatma(secili_model_listesi,satir,gorsel_klasor,gorsel_ismi):

                        def cikarim_model_liste_kontrol(secili_model_listesi,gecici_klasor_degisken,gorsel,satir):
                            if secili_model_listesi == []:
                                CTkMessagebox(title="Error", message="Prediction Failed to Start!!! \nModels Not Selected!!!",
                                              icon="cancel", width=400, height=200)
                            else:
                                self.kaydir = customtkinter.CTkScrollableFrame(self.cikarim, width=800, height=570)

                                self.kaydir.grid(column=10, row=0, pady=10, padx=10)
                                cikarim_algoritma_sonuc = cikarim_algoritma(secili_model_listesi,gecici_klasor_degisken,gorsel)

                                hesaplanan_veri = cikarim_algoritma_sonuc[0]
                                hesaplanan_sinir_agi = cikarim_algoritma_sonuc[1]
                                a = 0
                                b = 1

                                for grafik_veri in range(int(len(hesaplanan_veri) / 3)):
                                    wb = load_workbook(hesaplanan_veri[0 + a])  # processed excel file
                                    ws = wb.active
                                    toplam_satir = ws.max_row
                                    toplam_sutun = ws.max_column

                                    c = 0
                                    for sutun_islenmis in range(4, toplam_sutun + 1):
                                        satirlar_label = []
                                        degerler_label = []
                                        periyotlar = [0]
                                        # Reading the columns one by one
                                        string_deger_liste = []
                                        for satir_label in range(1, toplam_satir):
                                            if ws.cell(row=satir_label, column=sutun_islenmis).value == None:
                                                pass
                                            else:
                                                satirlar_label.append(satir_label)
                                                degerler_label.append(ws.cell(row=satir_label,
                                                                              column=sutun_islenmis).value)  # reading the lines

                                                if isinstance(ws.cell(row=satir_label + 1, column=sutun_islenmis).value,
                                                              str):
                                                    if ws.cell(row=satir_label + 1,
                                                               column=sutun_islenmis).value in string_deger_liste:
                                                        pass
                                                    else:
                                                        string_deger_liste.append(
                                                            ws.cell(row=satir_label + 1, column=sutun_islenmis).value)

                                            kirmizi = str(ws.cell(row=satir_label + 1,
                                                                  column=sutun_islenmis).font.color.__dict__).find(
                                                'FFFF0000')

                                            if kirmizi == -1:
                                                pass
                                            else:

                                                periyotlar.append(satir_label + 1)

                                        self.veri_goruntuleme = customtkinter.CTkFrame(self.kaydir)
                                        self.veri_goruntuleme.grid(row=b, column=0, columnspan=2)

                                        self.modelismi = customtkinter.CTkLabel(self.kaydir, text="{}".format(
                                            hesaplanan_veri[1 + a]))
                                        self.modelismi.grid(row=b - 1, column=0)

                                        self.goruntu_ismi = customtkinter.CTkLabel(self.kaydir, text="{}".format(
                                            hesaplanan_veri[0 + a]))
                                        self.goruntu_ismi.grid(row=b - 1, column=1)

                                        # AKAR Algorithm Prediction Results

                                        self.baslik = customtkinter.CTkTextbox(master=self.kaydir, width=800, height=30)
                                        self.baslik.grid(row=b + 1, column=0, columnspan=2, pady=2)
                                        self.baslik.insert("0.0", "AKAR Algorithm Prediction Results:")
                                        self.baslik.configure(state="disable")

                                        alt_veri_listesi = hesaplanan_veri[2 + a][sutun_islenmis - 4]

                                        try:
                                            if len(alt_veri_listesi) == 6:

                                                self.veri_seti_ismi = customtkinter.CTkTextbox(master=self.kaydir,
                                                                                               width=800, height=30)
                                                self.veri_seti_ismi.grid(row=b + 2, column=0, columnspan=2, pady=2)
                                                self.veri_seti_ismi.insert("0.0",
                                                                           "Comparison Performed Data Set:           {}".format(
                                                                               alt_veri_listesi[2]))
                                                self.veri_seti_ismi.configure(state="disable")

                                                self.veri_uyumu = customtkinter.CTkTextbox(master=self.kaydir, width=800,
                                                                                           height=30)
                                                self.veri_uyumu.grid(row=b + 3, column=0, columnspan=2, pady=2)
                                                self.veri_uyumu.insert("0.0",
                                                                       "Percentage of Fit to Model Data:          %{}".format(
                                                                           alt_veri_listesi[-1]))
                                                self.veri_uyumu.configure(state="disable")

                                                self.tahmini_deger_araligi = customtkinter.CTkTextbox(master=self.kaydir,
                                                                                                      width=800, height=30)
                                                self.tahmini_deger_araligi.grid(row=b + 4, column=0, columnspan=2, pady=2)
                                                self.tahmini_deger_araligi.insert("0.0",
                                                                                  "Estimated Value Range:         {}-{}".format(
                                                                                      alt_veri_listesi[3],
                                                                                      alt_veri_listesi[4]))
                                                self.tahmini_deger_araligi.configure(state="disable")

                                            elif len(alt_veri_listesi) == 2:
                                                self.veri_seti_ismi = customtkinter.CTkTextbox(master=self.kaydir,
                                                                                               width=800, height=30)
                                                self.veri_seti_ismi.grid(row=b + 2, column=0, columnspan=2, pady=2)
                                                self.veri_seti_ismi.insert("0.0",
                                                                           "Comparison Performed Data Set:           {}".format(
                                                                               alt_veri_listesi[0][2]))
                                                self.veri_seti_ismi.configure(state="disable")

                                                self.veri_uyumu1 = customtkinter.CTkTextbox(master=self.kaydir, width=800,
                                                                                            height=30)
                                                self.veri_uyumu1.grid(row=b + 3, column=0, columnspan=2, pady=2)
                                                self.veri_uyumu1.insert("0.0",
                                                                        "Percentage of Fit to Model Data 1:          %{}".format(
                                                                            alt_veri_listesi[0][-1]))
                                                self.veri_uyumu1.configure(state="disable")

                                                self.tahmini_deger_araligi1 = customtkinter.CTkTextbox(master=self.kaydir,
                                                                                                       width=800, height=30)
                                                self.tahmini_deger_araligi1.grid(row=b + 4, column=0, columnspan=2, pady=2)
                                                self.tahmini_deger_araligi1.insert("0.0",
                                                                                   "Estimated Value Range 1:         {}-{}".format(
                                                                                       alt_veri_listesi[0][3],
                                                                                       alt_veri_listesi[0][4]))
                                                self.tahmini_deger_araligi1.configure(state="disable")

                                                self.veri_uyumu2 = customtkinter.CTkTextbox(master=self.kaydir, width=800,
                                                                                            height=30)
                                                self.veri_uyumu2.grid(row=b + 5, column=0, columnspan=2, pady=2)
                                                self.veri_uyumu2.insert("0.0",
                                                                        "Percentage of Fit to Model Data 2:          %{}".format(
                                                                            alt_veri_listesi[1][-1]))
                                                self.veri_uyumu2.configure(state="disable")

                                                self.tahmini_deger_araligi2 = customtkinter.CTkTextbox(master=self.kaydir,
                                                                                                       width=800, height=30)
                                                self.tahmini_deger_araligi2.grid(row=b + 6, column=0, columnspan=2, pady=2)
                                                self.tahmini_deger_araligi2.insert("0.0",
                                                                                   "Estimated Value Range 2:         {}-{}".format(
                                                                                       alt_veri_listesi[1][3],
                                                                                       alt_veri_listesi[1][4]))
                                                self.tahmini_deger_araligi2.configure(state="disable")

                                        except:
                                            pass

                                        try:
                                            self.ara_cubuk = customtkinter.CTkLabel(master=self.kaydir, width=800,
                                                                                    height=10, text="", fg_color="grey70",
                                                                                    corner_radius=8)
                                            self.ara_cubuk.grid(row=b + 7, column=0, columnspan=2, pady=5)

                                            # Artificial Neural Network and Regression Prediction Results
                                            self.baslik = customtkinter.CTkTextbox(master=self.kaydir, width=800, height=30)
                                            self.baslik.grid(row=b + 8, column=0, columnspan=2, pady=2)
                                            self.baslik.insert("0.0", "Artificial Neural Network and Regression Prediction Results:")
                                            self.baslik.configure(state="disable")

                                            self.veri_seti_ismi2 = customtkinter.CTkTextbox(master=self.kaydir, width=800,
                                                                                            height=30)
                                            self.veri_seti_ismi2.grid(row=b + 9, column=0, columnspan=2, pady=2)

                                            self.veri_seti_ismi2.insert("0.0", "{}".format(
                                                hesaplanan_sinir_agi[grafik_veri][c + 1][
                                                    0]))
                                            self.veri_seti_ismi2.configure(state="disable")

                                            self.veri_uyumu2 = customtkinter.CTkTextbox(master=self.kaydir, width=800,
                                                                                        height=30)
                                            self.veri_uyumu2.grid(row=b + 10, column=0, columnspan=2, pady=2)
                                            self.veri_uyumu2.insert("0.0", "{}".format(
                                                hesaplanan_sinir_agi[grafik_veri][c + 1][
                                                    1]))
                                            self.veri_uyumu2.configure(state="disable")

                                            self.ara_cubuk2 = customtkinter.CTkLabel(master=self.kaydir, width=800,
                                                                                     height=2, text="", fg_color="grey70")
                                            self.ara_cubuk2.grid(row=b + 13, column=0, columnspan=2, pady=25)

                                            plt.style.use("cyberpunk")
                                            x_data = satirlar_label
                                            y_data = degerler_label
                                            fig = Figure(figsize=(10, 5), dpi=100)  # figure size

                                            ax1: Axes = fig.add_subplot(111)
                                            ax1.set_xlabel("Timeline")
                                            ax1.set_ylabel("{}".format(degerler_label[0]))
                                            satirlar_label.pop(0)
                                            degerler_label.pop(0)
                                            ax1.plot(x_data, y_data, linewidth=3, color='white')

                                            # Plotting neural network inference results into graphs
                                            # Plotting regression analysis results on a graph
                                            if isinstance(hesaplanan_sinir_agi[grafik_veri][c + 1][2], str):
                                                ax1.plot(hesaplanan_sinir_agi[grafik_veri][0][2], string_deger_liste.index(
                                                    hesaplanan_sinir_agi[grafik_veri][c + 1][2]), marker="o",
                                                         markersize=12, markerfacecolor='orange',
                                                         markeredgecolor="white",label="Artificial Neural Network and Regression Prediction Results")
                                                ax1.legend()
                                            else:
                                                ax1.plot(hesaplanan_sinir_agi[grafik_veri][0][2],
                                                         hesaplanan_sinir_agi[grafik_veri][c + 1][2], marker="o",
                                                         markersize=12, markerfacecolor='orange',
                                                         markeredgecolor="white",label="Artificial Neural Network and Regression Prediction Results")
                                                ax1.legend()
                                            # kontrol1 = If the list matches the none element, 2 lists are created, it tries by assuming 2 lists, if it gets an error, it enters the escape block.
                                            # kontrol2 = If the element is a string value, clustering is done in the list, each index in the list represents a different string expression.
                                            # Non-string elements are touched to the graph with X,Y coordinates
                                            try:
                                                tahmin_koordinat_x = hesaplanan_veri[2 + a][sutun_islenmis - 4][0][0]
                                                tahmin_koordinat_y = hesaplanan_veri[2 + a][sutun_islenmis - 4][0][1]

                                                # Prediction result
                                                tahmin_koordinat_x1 = hesaplanan_veri[2 + a][sutun_islenmis - 4][1][0]
                                                tahmin_koordinat_y1 = hesaplanan_veri[2 + a][sutun_islenmis - 4][1][1]

                                                if isinstance(tahmin_koordinat_y, str):
                                                    ax1.plot(tahmin_koordinat_x,
                                                             string_deger_liste.index(tahmin_koordinat_y), marker="o",
                                                             markersize=12, markerfacecolor='green',
                                                             markeredgecolor="white",label="Prediction of AKAR Algorithm")  # Prediction result
                                                    ax1.legend()


                                                else:
                                                    ax1.plot(tahmin_koordinat_x, tahmin_koordinat_y, marker="o",
                                                             markersize=12, markerfacecolor='green',
                                                             markeredgecolor="white",label="Prediction of AKAR Algorithm")
                                                    ax1.legend()


                                                if isinstance(tahmin_koordinat_y1, str):
                                                    ax1.plot(tahmin_koordinat_x1,
                                                             string_deger_liste.index(tahmin_koordinat_y1), marker="o",
                                                             markersize=12, markerfacecolor='red',
                                                             markeredgecolor="white",label="Prediction of AKAR Algorithm")  # Prediction result
                                                    ax1.legend()


                                                else:
                                                    ax1.plot(tahmin_koordinat_x1, tahmin_koordinat_y1, marker="o",
                                                             markersize=12, markerfacecolor='red', markeredgecolor="white",label="Prediction of AKAR Algorithm")
                                                    ax1.legend()

                                            except:
                                                tahmin_koordinat_x2 = hesaplanan_veri[2 + a][sutun_islenmis - 4][0]
                                                tahmin_koordinat_y2 = hesaplanan_veri[2 + a][sutun_islenmis - 4][1]

                                                if isinstance(tahmin_koordinat_y2, str):
                                                    ax1.plot(tahmin_koordinat_x2,
                                                             string_deger_liste.index(tahmin_koordinat_y2), marker="o",
                                                             markersize=12, markerfacecolor='green',
                                                             markeredgecolor="white",label="Prediction of AKAR Algorithm")  # Prediction result
                                                    ax1.legend()

                                                else:
                                                    ax1.plot(tahmin_koordinat_x2, tahmin_koordinat_y2, marker="o",
                                                             markersize=12, markerfacecolor='green',
                                                             markeredgecolor="white",label="Prediction of AKAR Algorithm")
                                                    ax1.legend()

                                            periyot = 0

                                            ax1.axvline(x=0, color='darkgrey',
                                                        label=' Periods')
                                            ax1.legend()



                                            for zaman in periyotlar:
                                                ax1.axvline(x=zaman, color='grey')
                                                periyot = periyot + 1

                                            ax1.set_title("Change Graph")

                                            canvas = FigureCanvasTkAgg(fig,
                                                                       master=self.veri_goruntuleme)
                                            canvas.draw()
                                            canvas.get_tk_widget().pack()
                                            toolbar = NavigationToolbar2Tk(canvas, self.veri_goruntuleme)
                                            toolbar.update()

                                            c = c + 1

                                            b = b + 17
                                        except:
                                            pass

                                    a = a + 3

                                def kontrol(satir):

                                    self.cikarim_baslat_label = customtkinter.CTkLabel(self.cikarim_veri_giris,
                                                                                       text="",
                                                                                       corner_radius=8, width=40,
                                                                                       fg_color="DarkOrange2", height=60)
                                    self.cikarim_baslat_label.grid(row=satir, column=3, pady=2)

                                threading.Thread(target=kontrol,args=str(satir)).start()

                            #Prediction start key
                        self.cikarim_baslatma_buton = customtkinter.CTkButton(self.cikarim_veri_giris,corner_radius=8, width=94, fg_color="green",height=60,text="Start\nPrediction",command=lambda gecici_klasor_degisken = gorsel_klasor,gorsel = gorsel_ismi: threading.Thread(target=cikarim_model_liste_kontrol,args=(secili_model_listesi,gecici_klasor_degisken,gorsel,satir,)).start())
                        self.cikarim_baslatma_buton.grid(row=satir, column=2, pady=2)

                    def secili_model_ekleme():
                        secili_model_listesi_demet = model_secme()

                        for ekle in secili_model_listesi_demet:
                            secili_model_listesi.append(ekle)

                    for gorseller in siralanmis_goruntu_klasor_icerigi:

                        self.cikarim_yonerge = customtkinter.CTkButton(self.cikarim_veri_giris, text="Directive",
                                                                           width=117, fg_color="red4",
                                                                           command=yonerge_cikarim_penceresi)
                        self.cikarim_yonerge.grid(row=0, column=0)

                        self.cikarim_icin_klasor = customtkinter.CTkButton(self.cikarim_veri_giris, text="Folder",
                                                                         width=82, fg_color="grey25",
                                                                         command=lambda kl = "klasor": threading.Thread(target=cikarim_goruntuler_al_klasor,args=(kl,)).start())
                        self.cikarim_icin_klasor.grid(row=0, column=1)

                        self.cikarim_icin_foto = customtkinter.CTkButton(self.cikarim_veri_giris, text="Image",
                                                                           width=82, fg_color="grey25",command=lambda tg = "tekil_gorsel" : cikarim_goruntuler_al_klasor(tg))
                        self.cikarim_icin_foto.grid(row=0, column=2)

                        #selection of models
                        self.model_sec = customtkinter.CTkButton(self.cikarim_veri_giris, text="Select Models",
                                                                 width=293,command=secili_model_ekleme)
                        self.model_sec.grid(row=1, column=0, columnspan=3,pady=5,padx=2)

                        self.numara1 = customtkinter.CTkLabel(self.cikarim_veri_giris, text="Image\nName",
                                                             corner_radius=8, width=120, fg_color="dark slate gray",
                                                             height=60)
                        self.numara1.grid(row=2, column=0, pady=2)

                        self.numara2 = customtkinter.CTkLabel(self.cikarim_veri_giris, text="Crop\nImage",
                                                              corner_radius=8, width=82, fg_color="DodgerBlue4",
                                                              height=60)
                        self.numara2.grid(row=2, column=1, pady=2)

                        self.numara3 = customtkinter.CTkLabel(self.cikarim_veri_giris, text="Crop\nControl",
                                                             corner_radius=8, width=82, fg_color="green", height=60)
                        self.numara3.grid(row=2, column=2, pady=2)

                        self.bilgi1 = customtkinter.CTkTextbox(master=self.cikarim_veri_giris,width=120,height=60)
                        self.bilgi1.grid(row=numara+2, column=0,pady=2)
                        self.bilgi1.insert("0.0","{}".format(liste_parcali[numara-1]))
                        self.bilgi1.configure(state="disable")

                        self.numara4 = customtkinter.CTkButton(self.cikarim_veri_giris, text="Crop",
                                                              corner_radius=8, width=82, fg_color="DodgerBlue4",height=60,command= lambda y = gorseller,satir = numara +2,foto_sadece_isim = liste_parcali[numara-1] :[koordinat_isaretleme(y,gecici_klasor),threading.Thread(target=cikarim_baslatma,args=(secili_model_listesi,satir,gecici_klasor, foto_sadece_isim,)).start()] )
                        self.numara4.grid(row=numara+2, column=1, pady=2)

                        self.numara5 = customtkinter.CTkLabel(self.cikarim_veri_giris, text="Image\nNot Cropped!",
                                                             corner_radius=8, width=82, fg_color="red4",height=60)
                        self.numara5.grid(row=numara+2, column=2, pady=2)

                        self.cikarim_baslat_label = customtkinter.CTkLabel(self.cikarim_veri_giris, text="",
                                                              corner_radius=8, width=40, fg_color="red4", height=60)
                        self.cikarim_baslat_label.grid(row=numara + 2, column=3, pady=2)

                        self.foto11 = customtkinter.CTkImage(
                            light_image=Image.open(Path(__file__).with_name("photos")/"m.png"), size=(25,110))

                        self.cikarim_baslat_kontrol = customtkinter.CTkLabel(self.cikarim_veri_giris , text="",image=self.foto11,
                                                                           corner_radius=8, fg_color="DarkOrange2",height=130,width=30)
                        self.cikarim_baslat_kontrol.grid(row=0,rowspan=3, column=3, pady=2)

                        numara = numara + 1

            def model_kiyaslama_cati():
                secili_modeller = model_secme()
                print(secili_modeller)
                model_compare.model_karsilastirma(secili_modeller)
                #transfer models to other function

            #first window
            self.cikarim_yonerge = customtkinter.CTkButton(self.cikarim_veri_giris, text="Directive", width=40,fg_color="red4",command=yonerge_cikarim_penceresi)
            self.cikarim_yonerge.grid(row=0, column=0,padx=2)


            self.cikarim_icin_klasor = customtkinter.CTkButton(self.cikarim_veri_giris,text="Folder",width=65,fg_color="grey25",command=lambda kl = "klasor": cikarim_goruntuler_al_klasor(kl))
            self.cikarim_icin_klasor.grid(row=0,column=1,padx=2)

            self.cikarim_icin_foto = customtkinter.CTkButton(self.cikarim_veri_giris, text="Image",width=65,fg_color="grey25",command=lambda tg = "tekil_gorsel" : cikarim_goruntuler_al_klasor(tg))
            self.cikarim_icin_foto.grid(row=0, column=2,padx=2)

            #Comparison of models
            self.model_karsilastirma = customtkinter.CTkButton(self.cikarim_veri_giris, text="Compare Models", width=50,command=model_kiyaslama_cati,
                                                             fg_color="grey25")
            self.model_karsilastirma.grid(row=0, column=3, padx=2)

            self.model_sec = customtkinter.CTkButton(self.cikarim_veri_giris,text="Select Models",width=330)
            self.model_sec.grid(row=1, column=0,columnspan=5,pady=5)


        #main window buttons
        self.button = customtkinter.CTkButton(self, command=egitim, width=600, text="Train", fg_color="green")
        self.button.grid(row=0, column=0, pady=5, padx=10)


        self.button = customtkinter.CTkButton(self, command=cikarim,width=600,text="Prediction")
        self.button.grid(row=0, column=1, pady=5, padx=10)


def butun_prosedurleri_sonlandirma():
    # Force kill all processes to kill CPU usage
    app.quit()
    app.destroy()

app = App()
app.protocol("WM_DELETE_WINDOW", butun_prosedurleri_sonlandirma)
app.mainloop()

