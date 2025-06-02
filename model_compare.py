import customtkinter as tk
import plotly.graph_objects as go
import webview
import openpyxl



def model_karsilastirma(model_listesi):
    ana_liste = []

    for model in model_listesi:
        ana_liste_alt = []
        liste_isim = []
        liste_isim.append(str(model))
        ana_liste_alt.append(liste_isim)  # adding model name

        wb_obj = openpyxl.load_workbook(model)
        sheet_obj = wb_obj.active
        row = sheet_obj.max_row
        column = sheet_obj.max_column

        for sutun in range(2, row + 1):  # Getting coordinates
            liste_bos = []
            cell_obj_x = sheet_obj.cell(row=sutun, column=1).value
            cell_obj_y = sheet_obj.cell(row=sutun, column=2).value
            cell_obj_z = sheet_obj.cell(row=sutun, column=3).value
            liste_bos.append(cell_obj_x)
            liste_bos.append(cell_obj_y)
            liste_bos.append(cell_obj_z)
            for satir in range(4, column + 1):
                cell_obj_label = sheet_obj.cell(row=1, column=satir).value
                cell_obj1 = sheet_obj.cell(row=sutun, column=satir).value
                liste_bos.append("{}:{}".format(cell_obj_label,cell_obj1))
            ana_liste_alt.append(liste_bos)
        ana_liste.append(ana_liste_alt)

    fig = go.Figure()

    colors = ['blue', 'red', 'green', 'purple', 'orange']

    for i, grup in enumerate(ana_liste):
        grup_adi = grup[0][0]  # First list element = group name
        x, y, z, text_labels = [], [], [], []

        for point in grup[1:]:  # Retrieve data other than group name
            x.append(point[0])
            y.append(point[1])
            z.append(point[2])
            text_labels.append(", ".join(map(str, point[3:])))  # Merge labels

        fig.add_trace(go.Scatter3d(
            x=x, y=y, z=z,
            mode='markers+text',
            marker=dict(size=4, color=colors[i % len(colors)], opacity=0.8),
            hovertext=text_labels,
            textposition="top center",
            name=f"{grup_adi} - Dots"
        ))

        fig.add_trace(go.Scatter3d(
            x=x, y=y, z=z,
            mode='lines',
            line=dict(color=colors[i % len(colors)], width=5),
            name=f"{grup_adi} - Line"
        ))

    fig.update_layout(
        scene_aspectmode='cube',

        paper_bgcolor='rgb(150, 150, 150)',
        plot_bgcolor='rgb(150, 150, 150)',
        font=dict(color='white'),
        scene=dict(


            xaxis=dict(range=[0, 255],color='black', backgroundcolor='rgb(70, 70, 70)',zerolinecolor="black",gridcolor="black",tick0= 15, dtick= 15,rangemode= 'tozero', tickmode= "linear",autorange=False),
            yaxis=dict(range=[0, 255],color='black', backgroundcolor='rgb(70, 70, 70)',zerolinecolor="black",gridcolor="black",tick0= 15, dtick= 15,rangemode= 'tozero', tickmode= "linear",autorange=False),
            zaxis=dict(range=[0, 255],color='black', backgroundcolor='rgb(70, 70, 70)',zerolinecolor="black",gridcolor="black",tick0= 15, dtick= 15,rangemode= 'tozero', tickmode= "linear",autorange=False),
            bgcolor='rgb(150, 150, 150)'
        )


    )

    fig.write_html("plotly_3d_dynamic_groups.html")


    webview.create_window('3D Plotly Dynamic Line Graph', 'plotly_3d_dynamic_groups.html', width=800, height=600)

    webview.start()

