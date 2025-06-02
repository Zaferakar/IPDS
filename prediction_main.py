from CTkMessagebox import CTkMessagebox
from PIL import Image
from openpyxl import load_workbook
import math
from softmax_random_forest_regression import *


def cikarim_algoritma(secili_model_listesi,gorsel_klasoru,gorsel):
    gorsel_konumu = "{}/{}".format(gorsel_klasoru,gorsel)
    def rgb_cikarici(gorsel_konumu):#rgb color value of the image

        im2 = Image.open(gorsel_konumu, 'r')
        pixel_values = list(im2.getdata())
        RGB_x = []
        RGB_y = []
        RGB_z = []
        for i in pixel_values:
            RGB_x.append(i[0])
            RGB_y.append(i[1])
            RGB_z.append(i[2])
        son_x = round(sum(RGB_x) / len(RGB_x))
        son_y = round(sum(RGB_y) / len(RGB_y))
        son_z = round(sum(RGB_z) / len(RGB_z))
        liste_bos = []
        liste_bos.append(son_x)
        liste_bos.append(son_y)
        liste_bos.append(son_z)

        return liste_bos#rgb code

    rgb_cikarim_degeri = rgb_cikarici(gorsel_konumu)

    butun_model_sonuclari = []
    sinir_agi_sonuc = []
    for excel_model in secili_model_listesi:
        butun_model_sonuclari.append(excel_model)
        butun_model_sonuclari.append(gorsel)

        wb = load_workbook(excel_model)
        ws = wb.active
        mesafe_listesi = []
        indeks_listesi = []
        def excel_oku_min_mesafe_hesapla():
            toplam_satir = ws.max_row  # total number of rows
            #First line X,Y,Z label
            for hucreler in range(2,toplam_satir+1):

                x_deger = int(ws.cell(row=hucreler, column=1).value)#X value
                y_deger = int(ws.cell(row=hucreler, column=2).value)#Y value
                z_deger = int(ws.cell(row=hucreler, column=3).value)#Z value

                fark1 = x_deger-rgb_cikarim_degeri[0]
                fark2 = y_deger-rgb_cikarim_degeri[1]
                fark3 = z_deger-rgb_cikarim_degeri[2]

                #distance calculation
                mesafe = ((fark1)**2+(fark2)**2+(fark3)**2)**0.5
                mesafe_listesi.append(mesafe)

            minimum_mesafe = min(mesafe_listesi)
            minimum_mesafe_indeks = mesafe_listesi.index(minimum_mesafe)
            indeks_listesi.append(minimum_mesafe_indeks)

            alinacak_satir_verileri = minimum_mesafe_indeks + 2

            satir_verileri_liste = []
            for hucre in ws[alinacak_satir_verileri]:
                satir_verileri_liste.append(hucre.value)

            return satir_verileri_liste,alinacak_satir_verileri#X,Y,Z + data

        karsilastirilacak_konum_ve_satir_bilgisi = excel_oku_min_mesafe_hesapla()#[X,Y,Z + data]

        karsilastirilacak_konum = karsilastirilacak_konum_ve_satir_bilgisi[0]
        def nokta_carpim_benzerligi_hesapla(karsilastirilacak_konum):
            #The starting point of the vector is assumed to be (0,0,0)
            karsilastirilacak_konum2 = [karsilastirilacak_konum[0],karsilastirilacak_konum[1],karsilastirilacak_konum[2]]
            nokta_carpimi = sum(a * b for a, b in zip(karsilastirilacak_konum2, rgb_cikarim_degeri))

            # Vector norms
            norm_a = math.sqrt(sum(a ** 2 for a in karsilastirilacak_konum2))
            norm_b = math.sqrt(sum(b ** 2 for b in rgb_cikarim_degeri))

            # Cosine similarity
            kosinus_benzerlik = nokta_carpimi / (norm_a * norm_b)

            # Norm ratio (size comparison)
            norm_orani = min(norm_a, norm_b) / max(norm_a, norm_b)
            yuzde_benzerlik = kosinus_benzerlik * norm_orani

            return yuzde_benzerlik

        nokta_carpim_benzerligi_islem_deger = nokta_carpim_benzerligi_hesapla(karsilastirilacak_konum)

        def verilerin_benzerlik_hesaplamasi(nokta_carpim_benzerligi_islem_deger):
            global sayisal_deger_sonuc,none_olasiligi_sonuc, string_olasilig_sonuc, sonuc_listesi
            sonuc_listesi = []
            for satir_degeri in range(3,len(karsilastirilacak_konum)):
                satir_veri = karsilastirilacak_konum[satir_degeri]

                #Calculation in case of None or NaN (inference based on the nearest control point)
                kolon_label_degeri = ws.cell(row=1, column=satir_degeri + 1).value
                def none_olasiligi(kolon_label_degeri):
                    global yollanacak_liste1, yollanacak_liste2
                    kolon_liste = []

                    for satirlar in range(1,ws.max_row):
                        satir_deger_none = ws.cell(row=satirlar+1, column=satir_degeri+1).value
                        kolon_liste.append(satir_deger_none)
                    indeks = indeks_listesi[0]

                    sayac1 = indeks
                    sayac2 = indeks
                    for secilmis_indeks in range(indeks):  # Scanning values up to the selected index
                        if kolon_liste[sayac1] == None:
                            sayac1 = sayac1 - 1
                        else:
                            break
                    for secilmis_indeks2 in range(
                            len(kolon_liste) - indeks):  # Scanning values from the selected index to the end of the list
                        if kolon_liste[sayac2] == None:
                            sayac2 = sayac2 + 1
                        else:
                            break
                    secili_nota_arasi_mesafe1 = indeks - sayac1
                    secili_nota_arasi_mesafe2 = sayac2 - indeks
                    if sayac1 + sayac2 == len(kolon_liste):
                        CTkMessagebox(title="Error", message="Error!!!",
                                      icon="cancel", width=400, height=200)
                    elif secili_nota_arasi_mesafe1 == secili_nota_arasi_mesafe2:
                        #Separate calculation of dot product value similarity
                        # the time dimension flowing backwards
                        xyz_yeni2 = []
                        for guncel_satir in range(0, 3):
                            yeni_satir = ws.cell(row=sayac1 + 2, column=guncel_satir + 1).value
                            xyz_yeni2.append(yeni_satir)
                        nokta_carpim_degeri_yeni1 = nokta_carpim_benzerligi_hesapla(
                            xyz_yeni2)  # It is assigned to the closest point and the cosine similarity is calculated
                        hucre_degeri1 = ws.cell(row=sayac1 + 2, column=satir_degeri + 1).value
                        # the forward flow of the time dimension
                        xyz_yeni = []
                        for guncel_satir in range(0, 3):
                            yeni_satir = ws.cell(row=sayac2 + 2, column=guncel_satir + 1).value
                            xyz_yeni.append(yeni_satir)

                        nokta_carpim_degeri_yeni2 = nokta_carpim_benzerligi_hesapla(
                            xyz_yeni)  # It is assigned to the closest point and the cosine similarity is calculated
                        hucre_degeri2 = ws.cell(row=sayac2 + 2, column=satir_degeri + 1).value

                        if isinstance(hucre_degeri2, str):
                            yollanacak_liste1 = [sayac1 + 2]#adding line number
                            yollanacak_liste2 = [sayac2 + 2]#adding line number
                            string_benzerlik1 = nokta_carpim_degeri_yeni1 * 100
                            string_benzerlik2 = nokta_carpim_degeri_yeni2*100
                            yollanacak_liste1.append(hucre_degeri1)
                            yollanacak_liste1.append(kolon_label_degeri)
                            yollanacak_liste1.append(hucre_degeri1)
                            yollanacak_liste1.append(hucre_degeri1)
                            yollanacak_liste1.append(string_benzerlik1)

                            yollanacak_liste2.append(hucre_degeri2)
                            yollanacak_liste2.append(kolon_label_degeri)
                            yollanacak_liste2.append(hucre_degeri2)
                            yollanacak_liste2.append(hucre_degeri2)
                            yollanacak_liste2.append(string_benzerlik2)
                        elif isinstance(hucre_degeri2, float) or isinstance(hucre_degeri2, int):
                            yollanacak_liste1 = [sayac1 + 2]#adding line number
                            yollanacak_liste2 = [sayac2 + 2]#adding line number
                            satir_veri_hesaplanmasi1 = hucre_degeri1 * (nokta_carpim_degeri_yeni1)
                            satir_veri_hesaplanmasi1_2 = hucre_degeri1 * (1-nokta_carpim_degeri_yeni1) + hucre_degeri1
                            satir_veri_hesaplanmasi2 = nokta_carpim_degeri_yeni2 * hucre_degeri2
                            satir_veri_hesaplanmasi2_2 = hucre_degeri2 * (1-nokta_carpim_degeri_yeni2)+hucre_degeri2
                            yollanacak_liste1.append(kolon_liste[sayac1])
                            yollanacak_liste1.append(kolon_label_degeri)
                            yollanacak_liste1.append(satir_veri_hesaplanmasi1)
                            yollanacak_liste1.append(satir_veri_hesaplanmasi1_2)
                            yollanacak_liste1.append(nokta_carpim_degeri_yeni1*100)

                            yollanacak_liste2.append(kolon_liste[sayac2])
                            yollanacak_liste2.append(kolon_label_degeri)
                            yollanacak_liste2.append(satir_veri_hesaplanmasi2)
                            yollanacak_liste2.append(satir_veri_hesaplanmasi2_2)
                            yollanacak_liste2.append(nokta_carpim_degeri_yeni2*100)
                        else:
                            pass

                        return yollanacak_liste1, yollanacak_liste2
                        #%50 %50
                    elif secili_nota_arasi_mesafe1 > secili_nota_arasi_mesafe2:
                        #reading new location
                        xyz_yeni = []
                        for guncel_satir in range(0,3):
                            yeni_satir = ws.cell(row=sayac2+2, column=guncel_satir + 1).value
                            xyz_yeni.append(yeni_satir)
                        nokta_carpim_degeri_yeni = nokta_carpim_benzerligi_hesapla(xyz_yeni)#assigned to the closest point and cosine similarity calculated
                        hucre_degeri = ws.cell(row=sayac2+2, column=satir_degeri + 1).value
                        if isinstance(hucre_degeri, str):

                            yollanacak_liste2 = [sayac2+2]
                            yollanacak_liste2.append(hucre_degeri)
                            yollanacak_liste2.append(kolon_label_degeri)
                            yollanacak_liste2.append(hucre_degeri)
                            yollanacak_liste2.append(hucre_degeri)
                            yollanacak_liste2.append(nokta_carpim_degeri_yeni*100)

                        elif isinstance(hucre_degeri, float) or isinstance(hucre_degeri, int):

                            yollanacak_liste2 = [sayac2+2]
                            benzerlik_orani = 100* nokta_carpim_degeri_yeni
                            satir_veri_hesaplanmasi = (nokta_carpim_degeri_yeni) * hucre_degeri
                            satir_veri_hesaplanmasi2 = (1-nokta_carpim_degeri_yeni) * hucre_degeri+hucre_degeri
                            yollanacak_liste2.append(hucre_degeri)
                            yollanacak_liste2.append(kolon_label_degeri)
                            yollanacak_liste2.append(satir_veri_hesaplanmasi)
                            yollanacak_liste2.append(satir_veri_hesaplanmasi2)
                            yollanacak_liste2.append(benzerlik_orani)
                        else:
                            pass

                        return yollanacak_liste2
                        #dot product similarity calculation
                    elif secili_nota_arasi_mesafe2 > secili_nota_arasi_mesafe1:
                        # reading new location
                        xyz_yeni = []
                        for guncel_satir in range(0, 3):
                            yeni_satir = ws.cell(row=sayac1 + 2, column=guncel_satir + 1).value
                            xyz_yeni.append(yeni_satir)
                        nokta_carpim_degeri_yeni = nokta_carpim_benzerligi_hesapla(
                            xyz_yeni)  # The closest point was assigned and the cosine similarity was calculated
                        hucre_degeri = ws.cell(row=sayac1 + 2, column=satir_degeri + 1).value
                        if isinstance(hucre_degeri, str):
                            yollanacak_liste1 = [sayac1 + 2]#adding line number
                            yollanacak_liste1.append(kolon_liste[sayac1])
                            yollanacak_liste1.append(hucre_degeri)
                            yollanacak_liste1.append(kolon_label_degeri)
                            yollanacak_liste1.append(hucre_degeri)
                            yollanacak_liste1.append((nokta_carpim_degeri_yeni)*100)
                        elif isinstance(hucre_degeri, float) or isinstance(hucre_degeri, int):
                            yollanacak_liste1 = [sayac1 + 2]#adding line number

                            benzerlik_orani = 100 * nokta_carpim_degeri_yeni
                            satir_veri_hesaplanmasi = (nokta_carpim_degeri_yeni) * hucre_degeri
                            satir_veri_hesaplanmasi2 = (1-nokta_carpim_degeri_yeni) * hucre_degeri+hucre_degeri
                            yollanacak_liste1.append(kolon_liste[sayac1])
                            yollanacak_liste1.append(kolon_label_degeri)
                            yollanacak_liste1.append(satir_veri_hesaplanmasi)
                            yollanacak_liste1.append(satir_veri_hesaplanmasi2)
                            yollanacak_liste1.append(benzerlik_orani)
                        else:
                            pass

                        return yollanacak_liste1
                    else:
                        pass
                    #Calculation in case of being a string (string expression + probability)

                def string_olasiligi(nokta_carpim_benzerligi_islem_deger,satir_veri,kolon_label_degeri):
                    yollanacak_liste1 = []
                    yollanacak_liste1.append(karsilastirilacak_konum_ve_satir_bilgisi[1])
                    yollanacak_liste1.append(satir_veri)
                    yollanacak_liste1.append(kolon_label_degeri)
                    yollanacak_liste1.append(satir_veri)
                    yollanacak_liste1.append(satir_veri)
                    yollanacak_liste1.append(nokta_carpim_benzerligi_islem_deger*100)

                    return yollanacak_liste1
                # Calculation in case of float and int(cosine similarity x value)

                def int_veya_float_olasiligi(nokta_carpim_benzerligi_islem_deger,hucre_degeri,kolon_label_degeri):
                    yollanacak_liste1 = []

                    benzerlik_orani = nokta_carpim_benzerligi_islem_deger*100
                    satir_veri_hesaplanmasi1 = (nokta_carpim_benzerligi_islem_deger)*hucre_degeri
                    satir_veri_hesaplanmasi2 = (1-nokta_carpim_benzerligi_islem_deger)*hucre_degeri+hucre_degeri
                    yollanacak_liste1.append(karsilastirilacak_konum_ve_satir_bilgisi[1])
                    yollanacak_liste1.append(hucre_degeri)
                    yollanacak_liste1.append(kolon_label_degeri)
                    yollanacak_liste1.append(satir_veri_hesaplanmasi1)
                    yollanacak_liste1.append(satir_veri_hesaplanmasi2)
                    yollanacak_liste1.append(benzerlik_orani)
                    return yollanacak_liste1
                if satir_veri is None:
                    none_olasiligi_sonuc = none_olasiligi(kolon_label_degeri)
                    sonuc_listesi.append(none_olasiligi_sonuc)
                elif isinstance(satir_veri, float) and math.isnan(satir_veri):  # NaN control
                    none_olasiligi_sonuc = none_olasiligi(kolon_label_degeri)
                    sonuc_listesi.append(none_olasiligi_sonuc)
                elif isinstance(satir_veri, str):
                    string_olasilig_sonuc = string_olasiligi(nokta_carpim_benzerligi_islem_deger,satir_veri,kolon_label_degeri)
                    sonuc_listesi.append(string_olasilig_sonuc)
                elif isinstance(satir_veri, float) or isinstance(satir_veri, int):
                    sayisal_deger_sonuc = int_veya_float_olasiligi(nokta_carpim_benzerligi_islem_deger,satir_veri,kolon_label_degeri)
                    sonuc_listesi.append(sayisal_deger_sonuc)
                else:
                    pass
            return sonuc_listesi

        veriler_son_deger = verilerin_benzerlik_hesaplamasi(nokta_carpim_benzerligi_islem_deger)
        butun_model_sonuclari.append(veriler_son_deger)


        #excel file(ws)
        toplam_satir = ws.max_row  # total number of rows
        toplam_sutun = ws.max_column  # total number of columns
        sinir_agi_besle_liste_xyz = []
        sinir_agi_zaman_cizgisi = []
        for hucreler in range(2, toplam_satir+1):
            liste_bos_gecici = []

            x_deger = int(ws.cell(row=hucreler, column=1).value)  # X value
            y_deger = int(ws.cell(row=hucreler, column=2).value)  # Y value
            z_deger = int(ws.cell(row=hucreler, column=3).value)  # Z value
            liste_bos_gecici.append(x_deger)
            liste_bos_gecici.append(y_deger)
            liste_bos_gecici.append(z_deger)
            sinir_agi_besle_liste_xyz.append(liste_bos_gecici)

            sinir_agi_zaman_cizgisi.append(hucreler-1)
        liste_sinir_agi_sonuclar = []
        koordinatxnoktasi = rastgele_orman_regresyonu(sinir_agi_besle_liste_xyz,sinir_agi_zaman_cizgisi,rgb_cikarim_degeri)
        liste_sinir_agi_sonuclar.append(koordinatxnoktasi)
        toplam_liste = []
        for sutun_ayri_oku in range(4, toplam_sutun + 1):
            sutun_veri = []
            sutun_veri_xyz = []
            label_degeri = ws.cell(row=1, column=sutun_ayri_oku).value
            toplam_liste.append(label_degeri)
            for sutun_oku in range(2, toplam_satir+1):
                satir_veri = ws.cell(row=sutun_oku, column=sutun_ayri_oku).value
                if satir_veri is None:
                    pass
                elif isinstance(satir_veri, float) or isinstance(satir_veri, int) or isinstance(satir_veri, str):
                    sutun_veri.append(satir_veri)

                    x_deger = int(ws.cell(row=sutun_oku, column=1).value)  # X value
                    y_deger = int(ws.cell(row=sutun_oku, column=2).value)  # Y value
                    z_deger = int(ws.cell(row=sutun_oku, column=3).value)   # Z value
                    koordinat = [x_deger,y_deger,z_deger]
                    sutun_veri_xyz.append(koordinat)
                else:
                    pass

            toplam_liste.append(sutun_veri_xyz)
            toplam_liste.append(sutun_veri)

        #change neural network according to data type
        #numerical data = regression
        #verbal data = SoftMax

        adim = 0
        for veri in range(int(len(toplam_liste)/3)):
            try:
                if isinstance(toplam_liste[adim+2][0], float) or isinstance(toplam_liste[adim+2][0], int):
                    liste_sinir_agi_sonuclar.append(rastgele_orman_regresyonu(toplam_liste[adim+1],toplam_liste[adim+2],rgb_cikarim_degeri))
                elif isinstance(toplam_liste[adim+2][0], str):
                    #convert data to tuple data type
                    veri_seti_sozluk = {}
                    indeks_deger = 0
                    for alt_listeler in toplam_liste[adim+1]:
                        alt_demet = tuple(alt_listeler)
                        veri_seti_sozluk[alt_demet] = toplam_liste[adim+2][indeks_deger]
                        indeks_deger = indeks_deger + 1

                    liste_sinir_agi_sonuclar.append(sinir_agi_softmax(veri_seti_sozluk,rgb_cikarim_degeri))
            except:
                pass
            adim = adim + 3

        sinir_agi_sonuc.append(liste_sinir_agi_sonuclar)

    return butun_model_sonuclari,sinir_agi_sonuc





