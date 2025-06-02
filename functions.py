import re
from tkinter import filedialog
from tkinter import *
import os
from natsort import natsorted
from openpyxl import Workbook,load_workbook
from CTkMessagebox import CTkMessagebox
from datetime import datetime

def model_secme():

    model_sec = filedialog.askopenfilenames()
    return model_sec

def gorsel_secme():
    gorsel_liste = []
    gorsel_sec = filedialog.askopenfilename()
    yol = os.path.dirname(gorsel_sec)

    gorsel_liste.append(gorsel_sec)
    gorsel_liste.append(yol)
    return gorsel_liste

def klasor_secme():
    ana_liste = []# Sorted location of images, folder location, excel file location,
    fotolar = []
    excel_dosyasi = []
    root = Tk()
    root.withdraw()
    folder_selected = filedialog.askdirectory()

    try:
        excel_sayac = 0

        dosyalar = os.listdir(folder_selected)
        for dosya in dosyalar:
            #adding excel files
            if dosya.endswith('.xlsx'):
                excel_dosyasi.append("{}/{}".format(folder_selected,dosya))
                excel_sayac = excel_sayac + 1
            elif dosya.endswith(".xls"):
                excel_dosyasi.append("{}/{}".format(folder_selected, dosya))
                excel_sayac = excel_sayac + 1
            #adding image files
            elif dosya.endswith(".png"):
                fotolar.append("{}/{}".format(folder_selected,dosya))
            elif dosya.endswith(".jpg"):
                fotolar.append("{}/{}".format(folder_selected, dosya))
            elif dosya.endswith(".jpeg"):
                fotolar.append("{}/{}".format(folder_selected, dosya))
            else:
                pass


        #If no photo is found in the folder, the user is warned
        if len(fotolar) == 0:
            CTkMessagebox(title="Error", message="No Image Found in Folder!!!",
                          icon="cancel", width=400, height=200)
        else:
            pass

        if excel_sayac == 1:
            pass
        elif excel_sayac == 0:
            CTkMessagebox(title="Error", message="Table Document Not Found in Folder!!!",
                          icon="cancel", width=400, height=200)
        elif excel_sayac >=2:
            CTkMessagebox(title="Error", message="More than One Table Document Found in Folder!!!", icon="cancel",width=400,height=200)
        else:
            pass

    except:
        pass

    siralanmis_klasor_icerigi = natsorted(fotolar)

    for foto_dosyalari in siralanmis_klasor_icerigi:
        ana_liste.append(foto_dosyalari)

    ana_liste.append(folder_selected)
    for excel_dosyasi_icin in excel_dosyasi:
        ana_liste.append(excel_dosyasi_icin)

    return ana_liste
#Reading Excel File

def excel_okuma(path,klasor_konum):
    foto_sayac = 0

    dosyalar = os.listdir(klasor_konum)
    for dosya in dosyalar:
        # adding image files
        if dosya.endswith(".png"):
            foto_sayac = foto_sayac + 1
        elif dosya.endswith(".jpg"):
            foto_sayac = foto_sayac + 1
        elif dosya.endswith(".jpeg"):
            foto_sayac = foto_sayac + 1
        else:
            pass

    wb = load_workbook(path)
    ws = wb.active

    sutun_no = 1
    sutun_boyut = 0


    while True:

        cell_obj = ws.cell(row=1, column=sutun_no)
        sutun_no= sutun_no+1

        if cell_obj.value == None:
            break
        else:
            sutun_boyut = sutun_boyut + 1

    sutun_no = 1
    ana_liste = []
    for i in range(sutun_boyut):
        sutun_liste = []

        satir_no = 1
        for i in range(foto_sayac+1):

            cell_obj = ws.cell(row=satir_no, column=sutun_no)

            satir_no = satir_no + 1
            if cell_obj.value == None:
                sutun_liste.append("None !!!")
            else:
                sutun_liste.append(cell_obj.value)

        ana_liste.append(sutun_liste)

        sutun_no = sutun_no + 1

    return sutun_boyut,foto_sayac+1,ana_liste

#creating an excel file for the analysis folder and data named analysis

def klasor_olusturma(klasor_konum):
    yil = datetime.now().year
    ay = datetime.now().month
    gun = datetime.now().day
    micro_saniye = datetime.now().microsecond

    zaman = "{}_{}_{}__{}".format(gun, ay, yil, micro_saniye)
    klasor_deger = "{}/{}".format(klasor_konum, zaman)
    os.mkdir(klasor_deger)

    return klasor_deger


